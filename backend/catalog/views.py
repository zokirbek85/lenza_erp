from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.files.base import ContentFile
from django.db import models
from django.http import FileResponse, HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated, AllowAny

from core.permissions import IsAccountant, IsAdmin, IsOwner, IsSales, IsWarehouse
from core.mixins.export_mixins import ExportMixin
from services.image_processing import process_product_image, delete_product_image, ImageProcessingError

from .models import Brand, Category, Product, Style
from .serializers import BrandSerializer, CategorySerializer, ProductSerializer, CatalogProductSerializer, StyleSerializer
from .utils.excel_tools import export_products_to_excel, export_products_to_excel_no_price, generate_import_template, import_products_from_excel


class BrandViewSet(viewsets.ModelViewSet):
    queryset = Brand.objects.all()
    serializer_class = BrandSerializer
    permission_classes = [IsAdmin | IsWarehouse | IsSales | IsAccountant | IsOwner]
    search_fields = ('name',)

    def get_queryset(self):
        queryset = super().get_queryset()
        # Brands are global - no dealer filtering needed
        return queryset


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAdmin | IsWarehouse | IsSales | IsAccountant | IsOwner]
    search_fields = ('name',)

    def get_queryset(self):
        queryset = super().get_queryset()
        brand_id = self.request.query_params.get('brand_id') or self.request.query_params.get('brand')
        if brand_id:
            queryset = queryset.filter(products__brand_id=brand_id).distinct()
        # Categories are global - no dealer filtering needed
        return queryset


class StyleViewSet(viewsets.ModelViewSet):
    queryset = Style.objects.filter(is_active=True).order_by('name')
    serializer_class = StyleSerializer
    permission_classes = [IsAuthenticated]
    search_fields = ('name',)

    def get_queryset(self):
        queryset = super().get_queryset()
        # Styles are global - no dealer filtering needed
        return queryset


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.select_related('category', 'brand').all()
    serializer_class = ProductSerializer
    permission_classes = [IsAdmin | IsAccountant | IsWarehouse | IsSales | IsOwner]
    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)
    filterset_fields = ('is_active', 'category', 'brand')
    search_fields = ('sku', 'name', 'barcode', 'article')
    ordering_fields = ('name', 'sell_price_usd', 'stock_ok', 'stock_defect')

    def get_queryset(self):
        queryset = super().get_queryset()
        query = self.request.query_params.get('q')
        if query:
            queryset = queryset.filter(models.Q(name__icontains=query) | models.Q(barcode__icontains=query))

        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(name__icontains=search)

        brand_id = self.request.query_params.get('brand_id') or self.request.query_params.get('brand')
        category_id = self.request.query_params.get('category_id') or self.request.query_params.get('category')
        style_id = self.request.query_params.get('style_id') or self.request.query_params.get('style')
        # Products are global - no dealer filtering
        if brand_id:
            queryset = queryset.filter(brand_id=brand_id)
        if category_id:
            queryset = queryset.filter(category_id=category_id)
        if style_id:
            queryset = queryset.filter(style_id=style_id)
        return queryset

    def list(self, request, *args, **kwargs):
        limit = request.query_params.get('limit')
        queryset = self.filter_queryset(self.get_queryset())
        if limit == 'all':
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
        return super().list(request, *args, **kwargs)

    def _ensure_manager(self):
        user = self.request.user
        if user.is_superuser:
            return
        if getattr(user, 'role', None) not in {'admin', 'owner'}:
            raise PermissionDenied('Only admin users can modify products.')

    def perform_create(self, serializer):
        self._ensure_manager()
        serializer.save()

    def perform_update(self, serializer):
        self._ensure_manager()
        serializer.save()

    def perform_destroy(self, instance):
        self._ensure_manager()
        instance.delete()

    @action(detail=True, methods=['patch'], url_path='adjust')
    def adjust_stock(self, request, pk=None):
        if not (request.user.is_superuser or getattr(request.user, 'role', None) == 'warehouse'):
            raise PermissionDenied('Only warehouse users can adjust stock levels.')
        product = self.get_object()
        stock_ok = request.data.get('stock_ok')
        stock_defect = request.data.get('stock_defect')
        updates = {}
        try:
            if stock_ok is not None:
                value = Decimal(str(stock_ok))
                updates['stock_ok'] = value.quantize(Decimal('0.01')) if value >= 0 else Decimal('0.00')
            if stock_defect is not None:
                value = Decimal(str(stock_defect))
                updates['stock_defect'] = value.quantize(Decimal('0.01')) if value >= 0 else Decimal('0.00')
        except (InvalidOperation, TypeError, ValueError):
            return Response({'detail': 'Stock values must be decimals.'}, status=400)
        if not updates:
            return Response({'detail': 'Provide stock_ok or stock_defect values.'}, status=400)
        Product.objects.filter(pk=product.pk).update(**updates)
        product.refresh_from_db()
        return Response(self.get_serializer(product).data)

    @action(detail=True, methods=['post'], url_path='upload-image', parser_classes=[MultiPartParser, FormParser])
    def upload_image(self, request, pk=None):
        """Upload and process product image with automatic background removal."""
        self._ensure_manager()
        product = self.get_object()
        
        # Validate file presence
        image_file = request.FILES.get('image')
        if not image_file:
            raise ValidationError({'image': 'Image file is required.'})
        
        # Validate file size (10 MB max)
        max_size = 10 * 1024 * 1024  # 10 MB
        if image_file.size > max_size:
            raise ValidationError({'image': 'Image file too large. Maximum size is 10 MB.'})
        
        # Validate file type
        allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/webp']
        content_type = getattr(image_file, 'content_type', '')
        if content_type not in allowed_types:
            raise ValidationError({'image': 'Invalid image format. Allowed: JPEG, PNG, WebP.'})
        
        try:
            # Delete old image if exists
            if product.image:
                old_path = product.image.name
                product.image.delete(save=False)
                delete_product_image(old_path)
            
            # Process image
            processed_bytes = process_product_image(
                image_file,
                max_size=1200,
                quality=85,
                remove_bg=True,
            )
            
            # Save processed image
            filename = f"{product.id}_{int(request.user.id or 0)}.webp"
            product.image.save(filename, ContentFile(processed_bytes), save=True)
            
            # Return updated product data
            serializer = self.get_serializer(product)
            return Response(serializer.data, status=status.HTTP_200_OK)
        
        except ImageProcessingError as e:
            raise ValidationError({'image': f'Image processing failed: {str(e)}'})
        except Exception as e:
            raise ValidationError({'image': f'Unexpected error: {str(e)}'})

    @action(detail=True, methods=['delete'], url_path='remove-image')
    def remove_image(self, request, pk=None):
        """Remove product image."""
        self._ensure_manager()
        product = self.get_object()
        
        if not product.image:
            return Response({'detail': 'No image to remove.'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            old_path = product.image.name
            product.image.delete(save=False)
            product.image = None
            product.save(update_fields=['image'])
            delete_product_image(old_path)
            
            serializer = self.get_serializer(product)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            raise ValidationError({'detail': f'Failed to remove image: {str(e)}'})


class DealerProductsAPIView(APIView):
    """
    Return all products with brand/category info.
    Products are global - not dealer-specific.
    This endpoint is kept for backward compatibility but returns all active products.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, dealer_id):
        # Products are global, return all active products
        products = (
            Product.objects.filter(is_active=True)
            .select_related('brand', 'category')
            .order_by('name')
        )
        data = [
            {
                'id': p.id,
                'name': p.name,
                'brand': {'id': p.brand.id, 'name': p.brand.name} if p.brand else None,
                'category': {'id': p.category.id, 'name': p.category.name} if p.category else None,
            }
            for p in products
        ]
        return Response(data)


class ProductExportExcelView(APIView):
    permission_classes = [IsAdmin | IsAccountant]

    def get(self, request):
        file_path = Path(export_products_to_excel())
        response = FileResponse(open(file_path, 'rb'), as_attachment=True, filename=file_path.name)
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response


class ProductImportExcelView(APIView):
    permission_classes = [IsAdmin]
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({'detail': 'Excel file is required.'}, status=status.HTTP_400_BAD_REQUEST)
        result = import_products_from_excel(excel_file)
        return Response(result, status=status.HTTP_201_CREATED)


class ProductImportTemplateView(APIView):
    permission_classes = [IsAdmin | IsAccountant]

    def get(self, request):
        file_path = Path(generate_import_template())
        response = FileResponse(open(file_path, 'rb'), as_attachment=True, filename=file_path.name)
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response


class ProductReportPDFView(APIView, ExportMixin):
    permission_classes = [IsAdmin | IsAccountant | IsSales | IsWarehouse]

    def get(self, request):
        products = Product.objects.select_related('brand', 'category').all()
        return self.render_pdf_with_qr(
            'reports/products_report.html',
            {'products': products},
            filename_prefix='products_report',
            request=request,
            doc_type='products-report',
            doc_id='bulk',
        )


class ProductCatalogNoPricePDFView(APIView):
    """Export products catalog as PDF without prices."""
    permission_classes = [IsAdmin | IsAccountant | IsSales | IsWarehouse | IsOwner]

    def get(self, request):
        from django.template.loader import render_to_string
        from weasyprint import HTML
        from django.utils import timezone
        
        products = Product.objects.select_related('brand', 'category').all()
        context = {
            'products': [
                {
                    'sku': p.sku,
                    'name': p.name,
                    'brand_name': p.brand.name if p.brand else '—',
                    'category_name': p.category.name if p.category else '—',
                    'stock_ok': float(p.stock_ok or 0),
                    'stock_defect': float(p.stock_defect or 0),
                }
                for p in products
            ],
            'export_date': timezone.now().strftime('%d.%m.%Y'),
        }
        
        html_content = render_to_string('catalog/catalog_no_price.html', context)
        pdf_file = HTML(string=html_content).write_pdf()
        
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="products_catalog_{timezone.now():%Y%m%d}.pdf"'
        return response


class ProductCatalogNoPriceExcelView(APIView):
    """Export products catalog as Excel without prices."""
    permission_classes = [IsAdmin | IsAccountant | IsSales | IsWarehouse | IsOwner]

    def get(self, request):
        file_path = Path(export_products_to_excel_no_price())
        response = FileResponse(open(file_path, 'rb'), as_attachment=True, filename=file_path.name)
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response


class CatalogView(APIView):
    """
    Catalog page API - returns door panel products grouped by brand with stock breakdown by width.
    Only shows products from category "Дверное полотно".
    """
    permission_classes = [IsAdmin | IsSales | IsAccountant | IsOwner]

    def get(self, request):
        # Filter products by category "Дверное полотно"
        products = Product.objects.filter(
            category__name='Дверное полотно',
            is_active=True
        ).select_related('brand', 'category').order_by('brand__name', 'name')
        
        serializer = CatalogProductSerializer(products, many=True, context={'request': request})
        return Response(serializer.data)


class CatalogExportPDFView(APIView, ExportMixin):
    """
    Export catalog as PDF with grouped products and komplektatsiya
    """
    permission_classes = [IsAdmin | IsSales | IsAccountant | IsOwner]

    def get(self, request):
        # Get filter parameters
        brand_filter = request.query_params.get('brand', 'all')
        search_query = request.query_params.get('search', '')
        view_mode = request.query_params.get('view', 'cards')  # cards, gallery

        # Get markup parameters
        markup_type = request.query_params.get('markup_type', 'percentage')  # 'percentage' or 'fixed'
        markup_value = request.query_params.get('markup_value', '0')

        try:
            markup_value = float(markup_value)
        except (ValueError, TypeError):
            markup_value = 0

        # Get all catalog variants (new variant-based system)
        variants = ProductVariant.objects.filter(is_active=True).select_related(
            'product_model__brand'
        ).prefetch_related('skus__product', 'kit_components__component')

        # Apply brand filter
        if brand_filter and brand_filter != 'all':
            variants = variants.filter(product_model__brand__name=brand_filter)

        # Order by brand and model
        variants = variants.order_by('product_model__brand__name', 'product_model__model_name', 'color')

        # Serialize variants with request context for absolute URLs
        serializer = VariantCatalogSerializer(variants, many=True, context={'request': request})
        product_data = list(serializer.data)  # Convert to list to ensure it's mutable

        # Convert image URLs to absolute file paths for PDF
        from django.conf import settings
        import os
        import logging

        logger = logging.getLogger(__name__)

        logger.info(f"PDF Export: Found {len(product_data)} products BEFORE image conversion")
        logger.info(f"PDF Export: Variants count from DB: {variants.count()}")

        for product in product_data:
            # Apply markup to prices if specified
            if markup_value > 0:
                if markup_type == 'percentage':
                    # Percentage markup
                    multiplier = 1 + (markup_value / 100)
                    if product.get('polotno_price_usd') is not None:
                        product['polotno_price_usd'] = round(product['polotno_price_usd'] * multiplier, 2)
                    if product.get('kit_price_usd') is not None:
                        product['kit_price_usd'] = round(product['kit_price_usd'] * multiplier, 2)
                    if product.get('full_set_price_usd') is not None:
                        product['full_set_price_usd'] = round(product['full_set_price_usd'] * multiplier, 2)
                else:
                    # Fixed amount markup
                    if product.get('polotno_price_usd') is not None:
                        product['polotno_price_usd'] = round(product['polotno_price_usd'] + markup_value, 2)
                    if product.get('kit_price_usd') is not None:
                        product['kit_price_usd'] = round(product['kit_price_usd'] + markup_value, 2)
                    if product.get('full_set_price_usd') is not None:
                        product['full_set_price_usd'] = round(product['full_set_price_usd'] + markup_value, 2)

            if product.get('image'):
                image_url = product['image']
                # Handle both relative and absolute URLs
                if '/media/' in image_url:
                    # Extract media path
                    media_path = image_url.split('/media/')[-1]
                    full_path = os.path.join(settings.MEDIA_ROOT, media_path)
                    full_path = os.path.normpath(full_path)  # Normalize path for Windows
                    if os.path.exists(full_path):
                        product['image'] = f'file:///{full_path}'  # Note: 3 slashes for Windows
                        logger.info(f"Image path converted: {image_url} -> {product['image']}")
                    else:
                        logger.warning(f"Image file not found: {full_path}")
                        product['image'] = None

        # Prepare context for template
        from datetime import date

        logger.info(f"PDF Export: Final product count: {len(product_data)}")
        if product_data:
            logger.info(f"First product keys: {list(product_data[0].keys())}")
            logger.info(f"First product data: {product_data[0]}")

        # Get logo path
        logo_path = None
        # Try multiple possible logo locations
        possible_logo_paths = [
            os.path.join(settings.BASE_DIR, 'frontend', 'public', 'logo-lenza-dark.png'),
            os.path.join(settings.BASE_DIR, 'frontend', 'public', 'logo.png'),
            os.path.join(settings.BASE_DIR, 'frontend', 'public', 'assets', 'logo-lenza-dark.png'),
            os.path.join(settings.BASE_DIR.parent, 'frontend', 'public', 'logo-lenza-dark.png'),
        ]
        
        for logo_file in possible_logo_paths:
            if os.path.exists(logo_file):
                logo_path = f'file:///{os.path.normpath(logo_file).replace(os.sep, "/")}'
                logger.info(f"Logo found: {logo_file}")
                logger.info(f"Logo URL: {logo_path}")
                break
        
        if not logo_path:
            logger.warning(f"Logo file not found. Tried: {possible_logo_paths}")
            # Try to get absolute path
            frontend_public = os.path.join(settings.BASE_DIR.parent, 'frontend', 'public')
            logger.warning(f"Frontend public dir exists: {os.path.exists(frontend_public)}")
            if os.path.exists(frontend_public):
                logger.warning(f"Files in public: {os.listdir(frontend_public)}")

        # Prepare markup display text
        markup_text = None
        if markup_value > 0:
            if markup_type == 'percentage':
                markup_text = f"+{markup_value}%"
            else:
                markup_text = f"+${markup_value}"

        context = {
            'products': product_data,
            'category': 'Дверное полотно',
            'brand_filter': brand_filter if brand_filter != 'all' else 'Все бренды',
            'view_mode': view_mode,
            'export_date': date.today().strftime('%d.%m.%Y'),
            'search_query': search_query,
            'total_products': len(product_data),
            'logo_path': logo_path,
            'markup_text': markup_text,
            'has_markup': markup_value > 0,
        }

        # Generate filename
        brand_slug = brand_filter.replace(' ', '_') if brand_filter != 'all' else 'all'
        filename_prefix = f'catalog_{brand_slug}'

        # Use test template for debugging
        return self.render_pdf_with_qr(
            'catalog/catalog_export_test.html',
            context,
            filename_prefix=filename_prefix,
            request=request,
            doc_type='catalog',
            doc_id='export',
        )


class CatalogExportExcelView(APIView):
    """
    Export catalog as Excel with grouped products
    """
    permission_classes = [IsAdmin | IsSales | IsAccountant | IsOwner]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from datetime import date
        from django.http import HttpResponse

        # Get filter parameters
        brand_filter = request.query_params.get('brand', 'all')
        search_query = request.query_params.get('search', '')

        # Get markup parameters
        markup_type = request.query_params.get('markup_type', 'percentage')  # 'percentage' or 'fixed'
        markup_value = request.query_params.get('markup_value', '0')

        try:
            markup_value = float(markup_value)
        except (ValueError, TypeError):
            markup_value = 0

        # Get all catalog products (using new variant-based system)
        variants = ProductVariant.objects.filter(is_active=True).select_related(
            'product_model__brand'
        ).prefetch_related('skus__product', 'kit_components__component')

        # Apply brand filter
        if brand_filter and brand_filter != 'all':
            variants = variants.filter(product_model__brand__name=brand_filter)

        # Order by brand and model
        variants = variants.order_by('product_model__brand__name', 'product_model__model_name', 'color')

        # Serialize variants
        serializer = VariantCatalogSerializer(variants, many=True, context={'request': request})
        product_data = list(serializer.data)

        # Apply markup to prices if specified
        if markup_value > 0:
            for product in product_data:
                if markup_type == 'percentage':
                    # Percentage markup
                    multiplier = 1 + (markup_value / 100)
                    if product.get('polotno_price_usd') is not None:
                        product['polotno_price_usd'] = round(product['polotno_price_usd'] * multiplier, 2)
                    if product.get('kit_price_usd') is not None:
                        product['kit_price_usd'] = round(product['kit_price_usd'] * multiplier, 2)
                    if product.get('full_set_price_usd') is not None:
                        product['full_set_price_usd'] = round(product['full_set_price_usd'] * multiplier, 2)
                else:
                    # Fixed amount markup
                    if product.get('polotno_price_usd') is not None:
                        product['polotno_price_usd'] = round(product['polotno_price_usd'] + markup_value, 2)
                    if product.get('kit_price_usd') is not None:
                        product['kit_price_usd'] = round(product['kit_price_usd'] + markup_value, 2)
                    if product.get('full_set_price_usd') is not None:
                        product['full_set_price_usd'] = round(product['full_set_price_usd'] + markup_value, 2)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Catalog'

        # Add markup info if present
        current_row = 1
        if markup_value > 0:
            markup_text = f"+{markup_value}%" if markup_type == 'percentage' else f"+${markup_value}"
            ws.merge_cells(f'A1:J1')
            info_cell = ws['A1']
            info_cell.value = f'Каталог с устамой (Markup): {markup_text}'
            info_cell.font = Font(bold=True, size=14, color='FF0000')
            info_cell.alignment = Alignment(horizontal='center', vertical='center')
            info_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            current_row = 2

        # Define headers
        headers = ['ID', 'Модель', 'Цвет', 'Бренд', 'Тип', 'Полотно ($)', 'Комплект ($)', 'Итого ($)', 'Макс. комплектов', 'Размеры в наличии']

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data
        row_num = current_row + 1
        for product in product_data:
            # Get sizes with stock
            sizes_with_stock = ', '.join([
                f"{s['size']}: {s['stock']}" for s in product.get('sizes', []) if s['stock'] > 0
            ]) or 'Н/Д'

            ws.cell(row=row_num, column=1, value=product.get('id'))
            ws.cell(row=row_num, column=2, value=product.get('model'))
            ws.cell(row=row_num, column=3, value=product.get('color'))
            ws.cell(row=row_num, column=4, value=product.get('brand'))
            ws.cell(row=row_num, column=5, value=product.get('door_type_display'))
            ws.cell(row=row_num, column=6, value=float(product.get('polotno_price_usd', 0)))
            ws.cell(row=row_num, column=7, value=float(product.get('kit_price_usd', 0)))
            ws.cell(row=row_num, column=8, value=float(product.get('full_set_price_usd', 0)))
            ws.cell(row=row_num, column=9, value=product.get('max_full_sets_by_stock') or 0)
            ws.cell(row=row_num, column=10, value=sizes_with_stock)
            row_num += 1

        # Auto-fit columns
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Generate response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        brand_slug = brand_filter.replace(' ', '_') if brand_filter != 'all' else 'all'
        today = date.today().strftime('%Y%m%d')
        filename = f'catalog_{brand_slug}_{today}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


# ========================================
# MARKETING DOCUMENT GENERATOR ENDPOINTS
# ========================================

class DealerCatalogPDFView(APIView, ExportMixin):
    """
    Generate dealer-specific catalog in PDF format with configurable markup and visibility options
    """
    permission_classes = [IsAdmin | IsSales | IsOwner]

    def get(self, request):
        from datetime import date
        from dealers.models import Dealer

        # Get parameters
        dealer_id = request.query_params.get('dealer_id')
        brand_filter = request.query_params.get('brand', 'all')
        markup_percent = float(request.query_params.get('markup', '0'))
        hide_price = request.query_params.get('hide_price', 'false') == 'true'
        hide_stock = request.query_params.get('hide_stock', 'false') == 'true'
        view_mode = request.query_params.get('view', 'cards')

        # Get dealer
        dealer = None
        if dealer_id:
            try:
                dealer = Dealer.objects.get(id=dealer_id)
            except Dealer.DoesNotExist:
                pass

        # Get products
        products = Product.objects.filter(
            category__name='Дверное полотно',
            is_active=True
        ).select_related('brand', 'category').order_by('brand__name', 'name')

        # Apply brand filter
        if brand_filter and brand_filter != 'all':
            products = products.filter(brand__name=brand_filter)

        # Serialize products with markup
        product_data = []
        for product in products:
            serializer = CatalogProductSerializer(product, context={'request': request})
            data = serializer.data
            
            # Apply markup to price
            if markup_percent > 0:
                original_price = float(data['price_usd'])
                data['price_usd'] = str(round(original_price * (1 + markup_percent / 100), 2))
            
            product_data.append(data)

        # Render PDF
        context = {
            'products': product_data,
            'brand_filter': brand_filter,
            'hide_price': hide_price,
            'hide_stock': hide_stock,
            'view_mode': view_mode,
            'dealer_name': dealer.name if dealer else 'Dealer',
            'markup_percent': markup_percent,
            'export_date': date.today().strftime('%d.%m.%Y'),
            'title': f'Каталог для дилера - {dealer.name}' if dealer else 'Каталог для дилера',
        }

        brand_slug = brand_filter.replace(' ', '_') if brand_filter != 'all' else 'all'
        filename = f'dealer_catalog_{dealer.id if dealer else "unknown"}_{brand_slug}_{date.today().strftime("%Y%m%d")}.pdf'

        return self.render_pdf_simple(
            template_path='marketing/dealer_catalog.html',
            context=context,
            filename=filename,
            request=request
        )


class DealerCatalogExcelView(APIView):
    """
    Generate dealer-specific catalog in Excel format with configurable markup and visibility options
    """
    permission_classes = [IsAdmin | IsSales | IsOwner]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        from datetime import date
        from django.http import HttpResponse
        from dealers.models import Dealer

        # Get parameters
        dealer_id = request.query_params.get('dealer_id')
        brand_filter = request.query_params.get('brand', 'all')
        markup_percent = float(request.query_params.get('markup', '0'))
        hide_price = request.query_params.get('hide_price', 'false') == 'true'
        hide_stock = request.query_params.get('hide_stock', 'false') == 'true'

        # Get dealer
        dealer = None
        if dealer_id:
            try:
                dealer = Dealer.objects.get(id=dealer_id)
            except Dealer.DoesNotExist:
                pass

        # Get products
        products = Product.objects.filter(
            category__name='Дверное полотно',
            is_active=True
        ).select_related('brand', 'category').order_by('brand__name', 'name')

        # Apply brand filter
        if brand_filter and brand_filter != 'all':
            products = products.filter(brand__name=brand_filter)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Dealer Catalog'

        # Define headers
        headers = ['Название', 'Бренд']
        if not hide_price:
            headers.append(f'Цена (USD){f" +{markup_percent}%" if markup_percent > 0 else ""}')
        if not hide_stock:
            headers.extend(['400мм', '600мм', '700мм', '800мм', '900мм', 'Всего'])

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data
        row_num = 2
        for product in products:
            serializer = CatalogProductSerializer(product, context={'request': request})
            data = serializer.data

            col_num = 1
            ws.cell(row=row_num, column=col_num, value=data['name'])
            col_num += 1
            ws.cell(row=row_num, column=col_num, value=data['brand_name'])
            col_num += 1

            if not hide_price:
                price = float(data['price_usd'])
                if markup_percent > 0:
                    price = price * (1 + markup_percent / 100)
                ws.cell(row=row_num, column=col_num, value=round(price, 2))
                col_num += 1

            if not hide_stock:
                stock = data.get('stock', {})
                stock_400 = stock.get('400', 0)
                stock_600 = stock.get('600', 0)
                stock_700 = stock.get('700', 0)
                stock_800 = stock.get('800', 0)
                stock_900 = stock.get('900', 0)
                total_stock = stock_400 + stock_600 + stock_700 + stock_800 + stock_900

                ws.cell(row=row_num, column=col_num, value=stock_400)
                col_num += 1
                ws.cell(row=row_num, column=col_num, value=stock_600)
                col_num += 1
                ws.cell(row=row_num, column=col_num, value=stock_700)
                col_num += 1
                ws.cell(row=row_num, column=col_num, value=stock_800)
                col_num += 1
                ws.cell(row=row_num, column=col_num, value=stock_900)
                col_num += 1
                ws.cell(row=row_num, column=col_num, value=total_stock)

            row_num += 1

        # Auto-fit columns
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Generate response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        dealer_name = dealer.name.replace(' ', '_') if dealer else 'unknown'
        brand_slug = brand_filter.replace(' ', '_') if brand_filter != 'all' else 'all'
        today = date.today().strftime('%Y%m%d')
        filename = f'dealer_catalog_{dealer_name}_{brand_slug}_{today}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


class BrandCatalogPDFView(APIView, ExportMixin):
    """
    Generate brand-specific catalog in PDF format
    """
    permission_classes = [IsAdmin | IsSales | IsOwner]

    def get(self, request):
        from datetime import date

        # Get parameters
        brand_name = request.query_params.get('brand')
        hide_price = request.query_params.get('hide_price', 'false') == 'true'
        hide_stock = request.query_params.get('hide_stock', 'false') == 'true'
        view_mode = request.query_params.get('view', 'cards')

        if not brand_name or brand_name == 'all':
            from django.http import JsonResponse
            return JsonResponse({'error': 'Brand name is required'}, status=400)

        # Get products for this brand only
        products = Product.objects.filter(
            category__name='Дверное полотно',
            brand__name=brand_name,
            is_active=True
        ).select_related('brand', 'category').order_by('name')

        # Serialize products
        serializer = CatalogProductSerializer(products, many=True, context={'request': request})
        product_data = serializer.data

        # Render PDF
        context = {
            'products': product_data,
            'brand_name': brand_name,
            'hide_price': hide_price,
            'hide_stock': hide_stock,
            'view_mode': view_mode,
            'export_date': date.today().strftime('%d.%m.%Y'),
            'title': f'Каталог бренда - {brand_name}',
        }

        brand_slug = brand_name.replace(' ', '_')
        filename = f'brand_catalog_{brand_slug}_{date.today().strftime("%Y%m%d")}.pdf'

        return self.render_pdf_simple(
            template_path='marketing/brand_catalog.html',
            context=context,
            filename=filename,
            request=request
        )


class BrandCatalogExcelView(APIView):
    """
    Generate brand-specific catalog in Excel format
    """
    permission_classes = [IsAdmin | IsSales | IsOwner]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        from datetime import date
        from django.http import HttpResponse, JsonResponse

        # Get parameters
        brand_name = request.query_params.get('brand')
        hide_price = request.query_params.get('hide_price', 'false') == 'true'
        hide_stock = request.query_params.get('hide_stock', 'false') == 'true'

        if not brand_name or brand_name == 'all':
            return JsonResponse({'error': 'Brand name is required'}, status=400)

        # Get products for this brand only
        products = Product.objects.filter(
            category__name='Дверное полотно',
            brand__name=brand_name,
            is_active=True
        ).select_related('brand', 'category').order_by('name')

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Brand Catalog'

        # Define headers
        headers = ['Название']
        if not hide_price:
            headers.append('Цена (USD)')
        if not hide_stock:
            headers.extend(['400мм', '600мм', '700мм', '800мм', '900мм', 'Всего'])

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data
        row_num = 2
        for product in products:
            serializer = CatalogProductSerializer(product, context={'request': request})
            data = serializer.data

            col_num = 1
            ws.cell(row=row_num, column=col_num, value=data['name'])
            col_num += 1

            if not hide_price:
                ws.cell(row=row_num, column=col_num, value=float(data['price_usd']))
                col_num += 1

            if not hide_stock:
                stock = data.get('stock', {})
                stock_400 = stock.get('400', 0)
                stock_600 = stock.get('600', 0)
                stock_700 = stock.get('700', 0)
                stock_800 = stock.get('800', 0)
                stock_900 = stock.get('900', 0)
                total_stock = stock_400 + stock_600 + stock_700 + stock_800 + stock_900

                ws.cell(row=row_num, column=col_num, value=stock_400)
                col_num += 1
                ws.cell(row=row_num, column=col_num, value=stock_600)
                col_num += 1
                ws.cell(row=row_num, column=col_num, value=stock_700)
                col_num += 1
                ws.cell(row=row_num, column=col_num, value=stock_800)
                col_num += 1
                ws.cell(row=row_num, column=col_num, value=stock_900)
                col_num += 1
                ws.cell(row=row_num, column=col_num, value=total_stock)

            row_num += 1

        # Auto-fit columns
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Generate response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        brand_slug = brand_name.replace(' ', '_')
        today = date.today().strftime('%Y%m%d')
        filename = f'brand_catalog_{brand_slug}_{today}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


class PriceListPDFView(APIView, ExportMixin):
    """
    Generate price list in PDF format (simplified, table-focused layout)
    """
    permission_classes = [IsAdmin | IsSales | IsOwner]

    def get(self, request):
        from datetime import date

        # Get parameters
        brand_filter = request.query_params.get('brand', 'all')
        hide_stock = request.query_params.get('hide_stock', 'false') == 'true'

        # Get products
        products = Product.objects.filter(
            category__name='Дверное полотно',
            is_active=True
        ).select_related('brand', 'category').order_by('brand__name', 'name')

        # Apply brand filter
        if brand_filter and brand_filter != 'all':
            products = products.filter(brand__name=brand_filter)

        # Serialize products
        serializer = CatalogProductSerializer(products, many=True, context={'request': request})
        product_data = serializer.data

        # Render PDF
        context = {
            'products': product_data,
            'brand_filter': brand_filter,
            'hide_stock': hide_stock,
            'export_date': date.today().strftime('%d.%m.%Y'),
            'title': f'Прайс-лист{f" - {brand_filter}" if brand_filter != "all" else ""}',
        }

        brand_slug = brand_filter.replace(' ', '_') if brand_filter != 'all' else 'all'
        filename = f'pricelist_{brand_slug}_{date.today().strftime("%Y%m%d")}.pdf'

        return self.render_pdf_simple(
            template_path='marketing/pricelist.html',
            context=context,
            filename=filename,
            request=request
        )


class PriceListExcelView(APIView):
    """
    Generate price list in Excel format (simplified, table-focused layout)
    """
    permission_classes = [IsAdmin | IsSales | IsOwner]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        from datetime import date
        from django.http import HttpResponse

        # Get parameters
        brand_filter = request.query_params.get('brand', 'all')
        hide_stock = request.query_params.get('hide_stock', 'false') == 'true'

        # Get products
        products = Product.objects.filter(
            category__name='Дверное полотно',
            is_active=True
        ).select_related('brand', 'category').order_by('brand__name', 'name')

        # Apply brand filter
        if brand_filter and brand_filter != 'all':
            products = products.filter(brand__name=brand_filter)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Price List'

        # Define headers
        headers = ['Название', 'Бренд', 'Цена (USD)']
        if not hide_stock:
            headers.extend(['400мм', '600мм', '700мм', '800мм', '900мм', 'Всего'])

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data
        row_num = 2
        for product in products:
            serializer = CatalogProductSerializer(product, context={'request': request})
            data = serializer.data

            ws.cell(row=row_num, column=1, value=data['name'])
            ws.cell(row=row_num, column=2, value=data['brand_name'])
            ws.cell(row=row_num, column=3, value=float(data['price_usd']))

            if not hide_stock:
                stock = data.get('stock', {})
                stock_400 = stock.get('400', 0)
                stock_600 = stock.get('600', 0)
                stock_700 = stock.get('700', 0)
                stock_800 = stock.get('800', 0)
                stock_900 = stock.get('900', 0)
                total_stock = stock_400 + stock_600 + stock_700 + stock_800 + stock_900

                ws.cell(row=row_num, column=4, value=stock_400)
                ws.cell(row=row_num, column=5, value=stock_600)
                ws.cell(row=row_num, column=6, value=stock_700)
                ws.cell(row=row_num, column=7, value=stock_800)
                ws.cell(row=row_num, column=8, value=stock_900)
                ws.cell(row=row_num, column=9, value=total_stock)

            row_num += 1

        # Auto-fit columns
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Generate response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        brand_slug = brand_filter.replace(' ', '_') if brand_filter != 'all' else 'all'
        today = date.today().strftime('%Y%m%d')
        filename = f'pricelist_{brand_slug}_{today}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


# ============================================================================
# VARIANT-BASED CATALOG VIEWS (for "Дверное полотно" category)
# ============================================================================

from .models import ProductModel, ProductVariant, ProductSKU
from .serializers import (
    ProductModelSerializer,
    ProductVariantDetailSerializer,
    ProductSKUSerializer,
    VariantCatalogSerializer,
    PublicVariantCatalogSerializer,
)


class ProductModelViewSet(viewsets.ModelViewSet):
    """ViewSet for product models"""
    queryset = ProductModel.objects.filter(is_active=True).select_related('brand')
    serializer_class = ProductModelSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ('brand', 'collection')
    search_fields = ('model_name', 'collection', 'brand__name')
    ordering_fields = ('brand__name', 'collection', 'model_name')
    ordering = ('brand__name', 'collection', 'model_name')


class ProductVariantViewSet(viewsets.ModelViewSet):
    """ViewSet for product variants - Admin only for write operations"""
    queryset = ProductVariant.objects.all().select_related('product_model__brand')
    serializer_class = ProductVariantDetailSerializer
    permission_classes = [IsAdmin | IsOwner]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ('product_model', 'door_type', 'is_active')
    search_fields = ('product_model__model_name', 'color', 'product_model__brand__name')
    ordering_fields = ('product_model__brand__name', 'product_model__model_name', 'color')
    ordering = ('product_model__brand__name', 'product_model__model_name', 'color')
    parser_classes = [MultiPartParser, FormParser]
    
    def create(self, request, *args, **kwargs):
        """Create a new product variant with proper error handling"""
        import logging
        logger = logging.getLogger(__name__)
        
        try:
            logger.info(f"Creating variant with data: {request.data}")
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
        except ValidationError as e:
            logger.error(f"Validation error creating variant: {e.detail}")
            return Response(
                {'detail': 'Validation error', 'errors': e.detail},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.exception(f"Exception creating variant: {str(e)}")
            import traceback
            return Response(
                {'detail': f'Error creating variant: {str(e)}', 'traceback': traceback.format_exc()},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def update(self, request, *args, **kwargs):
        """Update a product variant with proper error handling"""
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data)
        except ValidationError as e:
            return Response(
                {'detail': 'Validation error', 'errors': e.detail},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'detail': f'Error updating variant: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ProductSKUViewSet(viewsets.ModelViewSet):
    """ViewSet for product SKUs"""
    queryset = ProductSKU.objects.all().select_related(
        'variant__product_model__brand',
        'product'
    )
    serializer_class = ProductSKUSerializer
    permission_classes = [IsAdmin | IsWarehouse | IsSales | IsAccountant | IsOwner]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ('variant', 'size')
    search_fields = ('product__sku', 'product__name')
    ordering_fields = ('variant', 'size')
    ordering = ('variant', 'size')


class VariantCatalogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Main API endpoint for variant-based catalog.
    Returns ProductVariant as card unit with size/stock breakdown.
    
    Filters:
    - brand: Filter by brand name (partial match)
    - model: Filter by model name (partial match)
    - color: Filter by color (partial match)
    - door_type: Filter by door type (ПГ, ПО, ПДО, ПДГ)
    - size: Filter by available size
    - search: Search in model name, color, brand
    
    Example:
    GET /api/catalog/variants/?brand=ДУБРАВА&door_type=ПГ&size=800мм
    """
    queryset = ProductVariant.objects.filter(is_active=True).select_related(
        'product_model__brand'
    ).prefetch_related('skus__product', 'kit_components__component')
    serializer_class = VariantCatalogSerializer
    permission_classes = [AllowAny]  # Katalog umumiy (public)
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ('product_model__model_name', 'color', 'product_model__brand__name', 'product_model__collection')
    ordering_fields = ('product_model__brand__name', 'product_model__model_name', 'color')
    ordering = ('product_model__brand__name', 'product_model__model_name', 'color')
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by brand name (partial match)
        brand = self.request.query_params.get('brand')
        if brand:
            queryset = queryset.filter(product_model__brand__name__icontains=brand)
        
        # Filter by model name (partial match)
        model = self.request.query_params.get('model')
        if model:
            queryset = queryset.filter(product_model__model_name__icontains=model)
        
        # Filter by collection
        collection = self.request.query_params.get('collection')
        if collection:
            queryset = queryset.filter(product_model__collection__icontains=collection)
        
        # Filter by color (partial match)
        color = self.request.query_params.get('color')
        if color:
            queryset = queryset.filter(color__icontains=color)
        
        # Filter by door type
        door_type = self.request.query_params.get('door_type')
        if door_type:
            queryset = queryset.filter(door_type=door_type)
        
        # Filter by size (variants that have SKUs with this size)
        size = self.request.query_params.get('size')
        if size:
            queryset = queryset.filter(skus__size=size).distinct()
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def filters(self, request):
        """
        Return available filter options from active variants
        """
        variants = ProductVariant.objects.filter(is_active=True).select_related('product_model__brand')
        
        # Extract unique values
        brands = set()
        collections = set()
        models = set()
        colors = set()
        door_types = set()
        
        for variant in variants:
            brands.add(variant.product_model.brand.name)
            collections.add(variant.product_model.collection)
            models.add(variant.product_model.model_name)
            colors.add(variant.color)
            door_types.add((variant.door_type, variant.get_door_type_display()))
        
        return Response({
            'brands': sorted(list(brands)),
            'collections': sorted(list(collections)),
            'models': sorted(list(models)),
            'colors': sorted(list(colors)),
            'door_types': [{'value': val, 'label': label} for val, label in sorted(door_types, key=lambda x: x[1])],
            'sizes': [
                {'value': '400мм', 'label': '400мм'},
                {'value': '600мм', 'label': '600мм'},
                {'value': '700мм', 'label': '700мм'},
                {'value': '800мм', 'label': '800мм'},
                {'value': '900мм', 'label': '900мм'},
            ]
        })


# ============================================================================
# PUBLIC CATALOG API (for external customers)
# ============================================================================

class PublicVariantCatalogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Public catalog API for external customers.
    
    Available at: /api/public/catalog/variants/
    
    Features:
    - No authentication required
    - No prices shown
    - No stock quantities shown
    - Only active products with images
    - Fully responsive for mobile devices
    
    Used for public catalogue page at https://lenza.uz/catalogue
    
    Query Parameters:
    - brand: Filter by brand name (partial match)
    - model: Filter by model name (partial match)
    - color: Filter by color (partial match)
    - door_type: Filter by door type (ПГ, ПО, ПДО, ПДГ)
    - search: Search in model name, color, brand
    """
    serializer_class = PublicVariantCatalogSerializer
    permission_classes = []  # No authentication required
    authentication_classes = []  # No authentication
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ('product_model__model_name', 'color', 'product_model__brand__name')
    ordering_fields = ('product_model__brand__name', 'product_model__model_name', 'color')
    ordering = ('product_model__brand__name', 'product_model__model_name', 'color')
    
    def get_queryset(self):
        """
        Return only active variants with images.
        Public customers should only see products with images.
        """
        queryset = ProductVariant.objects.filter(
            is_active=True,
            image__isnull=False  # Only variants with images
        ).exclude(
            image=''  # Exclude empty image paths
        ).select_related(
            'product_model__brand'
        ).prefetch_related(
            'skus__product'
        )
        
        # Filter by brand name (partial match)
        brand = self.request.query_params.get('brand')
        if brand:
            queryset = queryset.filter(product_model__brand__name__icontains=brand)
        
        # Filter by model name (partial match)
        model = self.request.query_params.get('model')
        if model:
            queryset = queryset.filter(product_model__model_name__icontains=model)
        
        # Filter by color (partial match)
        color = self.request.query_params.get('color')
        if color:
            queryset = queryset.filter(color__icontains=color)
        
        # Filter by door type (exact match)
        door_type = self.request.query_params.get('door_type')
        if door_type:
            queryset = queryset.filter(door_type=door_type)
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def filter_options(self, request):
        """
        Return available filter options for public catalog.
        Useful for filter dropdowns in frontend.
        """
        queryset = self.get_queryset()
        
        brands = set()
        models = set()
        colors = set()
        door_types = set()
        
        for variant in queryset:
            brands.add(variant.product_model.brand.name)
            models.add(variant.product_model.model_name)
            colors.add(variant.color)
            door_types.add((variant.door_type, variant.get_door_type_display()))
        
        return Response({
            'brands': sorted(list(brands)),
            'models': sorted(list(models)),
            'colors': sorted(list(colors)),
            'door_types': [{'value': val, 'label': label} for val, label in sorted(door_types, key=lambda x: x[1])],
        })
