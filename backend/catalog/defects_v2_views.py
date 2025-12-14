"""
Defects V2 Views
================

DRF ViewSets for the new defects module.
Provides REST API endpoints with proper permissions and business logic.
"""

from decimal import Decimal
from datetime import datetime
from django.db.models import Q, Sum, Count
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from core.permissions import IsAdmin, IsWarehouse
from .defects_v2_models import (
    DefectType,
    SparePartV2,
    DefectBatchV2,
    DefectDetailV2,
    DefectRepairV2,
    RepairMaterialV2,
    DefectWriteOffV2,
    DefectAuditLogV2,
)
from .defects_v2_serializers import (
    DefectTypeSerializer,
    SparePartSerializer,
    DefectBatchListSerializer,
    DefectBatchDetailSerializer,
    CreateDefectBatchSerializer,
    InspectBatchSerializer,
    DefectRepairSerializer,
    RepairDefectSerializer,
    DefectWriteOffSerializer,
    WriteOffDefectSerializer,
    DefectAuditLogSerializer,
    DefectStatisticsSerializer,
)
from .defects_service import DefectService, DefectAnalyticsService


# ============================================================================
# Reference Data ViewSets
# ============================================================================

class DefectTypeViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing defect types reference table.

    Available actions:
    - list: Get all defect types
    - retrieve: Get single defect type
    - create: Create new defect type (admin only)
    - update: Update defect type (admin only)
    - destroy: Delete defect type (admin only)
    """

    queryset = DefectType.objects.all()
    serializer_class = DefectTypeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'is_active']
    search_fields = ['name', 'name_uz', 'name_en', 'description']
    ordering_fields = ['name', 'category', 'created_at']
    ordering = ['name']

    def get_permissions(self):
        """Only admin can create/update/delete defect types"""
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAuthenticated(), IsAdmin()]
        return [IsAuthenticated()]


class SparePartViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing spare parts.

    Available actions:
    - list: Get all spare parts
    - retrieve: Get single spare part with stock info
    - create: Create new spare part (admin/warehouse)
    - update: Update spare part (admin/warehouse)
    - destroy: Delete spare part (admin only)
    - low_stock: Get spare parts below minimum stock level
    """

    queryset = SparePartV2.objects.select_related('product').all()
    serializer_class = SparePartSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['is_active']
    search_fields = ['name', 'product__name', 'product__sku']
    ordering_fields = ['name', 'min_stock_level', 'created_at']
    ordering = ['name']

    def get_permissions(self):
        """Admin and warehouse can create/update, only admin can delete"""
        if self.action == 'destroy':
            return [IsAuthenticated(), IsAdmin()]
        elif self.action in ['create', 'update', 'partial_update']:
            return [IsAuthenticated(), (IsAdmin() | IsWarehouse())]
        return [IsAuthenticated()]

    @action(detail=False, methods=['get'], url_path='low-stock')
    def low_stock(self, request):
        """Get spare parts with stock below minimum level"""
        low_stock_parts = [
            sp for sp in self.get_queryset()
            if sp.is_low_stock()
        ]
        serializer = self.get_serializer(low_stock_parts, many=True)
        return Response(serializer.data)


# ============================================================================
# Defect Batch ViewSet
# ============================================================================

class DefectBatchViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing defect batches.

    Available actions:
    - list: Get all batches (lightweight)
    - retrieve: Get single batch with full details
    - create: Create new batch
    - update: Update batch metadata (notes, warehouse)
    - inspect: Mark batch as inspected with defect details
    - repair: Perform repair operation
    - write_off: Perform write-off operation
    - complete: Mark batch as completed
    - audit_log: Get audit trail for a batch
    """

    queryset = DefectBatchV2.objects.select_related(
        'product', 'created_by'
    ).prefetch_related('defect_details__defect_type').all()
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'product']
    search_fields = ['product__name', 'product__sku', 'notes']
    ordering_fields = ['detected_at', 'total_qty', 'status']
    ordering = ['-detected_at']

    def get_serializer_class(self):
        """Use lightweight serializer for list, detailed for retrieve"""
        if self.action == 'list':
            return DefectBatchListSerializer
        elif self.action == 'create':
            return CreateDefectBatchSerializer
        elif self.action == 'inspect':
            return InspectBatchSerializer
        elif self.action == 'repair':
            return RepairDefectSerializer
        elif self.action == 'write_off':
            return WriteOffDefectSerializer
        return DefectBatchDetailSerializer

    def get_queryset(self):
        """Apply additional filters"""
        queryset = super().get_queryset()

        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')

        if start_date:
            try:
                start = datetime.fromisoformat(start_date)
                queryset = queryset.filter(detected_at__gte=start)
            except ValueError:
                pass

        if end_date:
            try:
                end = datetime.fromisoformat(end_date)
                queryset = queryset.filter(detected_at__lte=end)
            except ValueError:
                pass

        return queryset

    def create(self, request, *args, **kwargs):
        """Create new defect batch"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Use service layer to create batch
        batch = DefectService.create_defect_batch(
            product_id=serializer.validated_data['product_id'],
            total_qty=serializer.validated_data['total_qty'],
            repairable_qty=serializer.validated_data['repairable_qty'],
            non_repairable_qty=serializer.validated_data['non_repairable_qty'],
            defect_details=serializer.validated_data.get('defect_details', []),
            warehouse_id=serializer.validated_data.get('warehouse_name', ''),
            notes=serializer.validated_data.get('notes', ''),
            user=request.user
        )

        # Return detailed representation
        output_serializer = DefectBatchDetailSerializer(batch)
        return Response(output_serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def inspect(self, request, pk=None):
        """
        Mark batch as inspected with defect details.

        POST /api/defects/batches/{id}/inspect/
        {
            "repairable_qty": 15.00,
            "non_repairable_qty": 5.00,
            "defect_details": [
                {"defect_type_id": 1, "qty": 3, "notes": "..."},
                {"defect_type_id": 2, "qty": 2}
            ]
        }
        """
        batch = self.get_object()
        serializer = self.get_serializer(data=request.data, context={'batch': batch})
        serializer.is_valid(raise_exception=True)

        # Use service layer to inspect
        batch = DefectService.inspect_batch(
            batch_id=batch.id,
            repairable_qty=serializer.validated_data['repairable_qty'],
            non_repairable_qty=serializer.validated_data['non_repairable_qty'],
            defect_details=serializer.validated_data['defect_details'],
            user=request.user
        )

        output_serializer = DefectBatchDetailSerializer(batch)
        return Response(output_serializer.data)

    @action(detail=True, methods=['post'])
    def repair(self, request, pk=None):
        """
        Perform repair operation on batch.

        POST /api/defects/batches/{id}/repair/
        {
            "qty": 10.00,
            "spare_parts": [
                {"spare_part_id": 1, "qty": 5, "unit_cost_usd": 2.50},
                {"spare_part_id": 2, "qty": 10, "unit_cost_usd": 1.00}
            ],
            "notes": "Replaced damaged parts"
        }
        """
        batch = self.get_object()
        serializer = self.get_serializer(data=request.data, context={'batch': batch})
        serializer.is_valid(raise_exception=True)

        # Use service layer to repair
        repair = DefectService.repair_defect(
            batch_id=batch.id,
            qty=serializer.validated_data['qty'],
            spare_parts=serializer.validated_data.get('spare_parts', []),
            notes=serializer.validated_data.get('notes', ''),
            user=request.user
        )

        output_serializer = DefectRepairSerializer(repair)
        return Response(output_serializer.data)

    @action(detail=True, methods=['post'])
    def write_off(self, request, pk=None):
        """
        Perform write-off operation on batch.

        POST /api/defects/batches/{id}/write_off/
        {
            "qty": 5.00,
            "reason": "disposal",
            "notes": "Beyond repair",
            "sale_price_usd": null
        }
        """
        batch = self.get_object()
        serializer = self.get_serializer(data=request.data, context={'batch': batch})
        serializer.is_valid(raise_exception=True)

        # Use service layer to write off
        write_off = DefectService.write_off_defect(
            batch_id=batch.id,
            qty=serializer.validated_data['qty'],
            reason=serializer.validated_data['reason'],
            notes=serializer.validated_data.get('notes', ''),
            sale_price_usd=serializer.validated_data.get('sale_price_usd'),
            user=request.user
        )

        output_serializer = DefectWriteOffSerializer(write_off)
        return Response(output_serializer.data)

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """
        Mark batch as completed.

        POST /api/defects/batches/{id}/complete/
        """
        batch = self.get_object()

        # Check if fully processed
        if not batch.is_fully_processed():
            return Response(
                {
                    'detail': f'Batch not fully processed. Remaining: {batch.remaining_qty()} units. '
                              f'Repairable: {batch.repairable_qty}, Non-repairable: {batch.non_repairable_qty}'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # Update status
        batch.status = DefectBatchV2.Status.COMPLETED
        batch.completed_at = timezone.now()
        batch.save()

        # Create audit log
        DefectAuditLogV2.objects.create(
            batch=batch,
            action=DefectAuditLogV2.Action.STATUS_CHANGED,
            performed_by=request.user,
            old_data={'status': 'processing'},
            new_data={'status': 'completed'},
            description=f'Batch marked as completed by {request.user.get_full_name()}'
        )

        output_serializer = DefectBatchDetailSerializer(batch)
        return Response(output_serializer.data)

    @action(detail=True, methods=['get'], url_path='audit-log')
    def audit_log(self, request, pk=None):
        """Get audit trail for a batch"""
        batch = self.get_object()
        logs = DefectAuditLogV2.objects.filter(batch=batch).order_by('-performed_at')
        serializer = DefectAuditLogSerializer(logs, many=True)
        return Response(serializer.data)


# ============================================================================
# Repair and Write-Off ViewSets
# ============================================================================

class DefectRepairViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for viewing repair records.
    Repairs are created through DefectBatchViewSet.repair action.
    """

    queryset = DefectRepairV2.objects.select_related(
        'batch__product', 'performed_by'
    ).prefetch_related('materials__spare_part').all()
    serializer_class = DefectRepairSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['batch', 'status', 'performed_by']
    ordering_fields = ['started_at', 'completed_at', 'qty_repaired']
    ordering = ['-started_at']


class DefectWriteOffViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for viewing write-off records.
    Write-offs are created through DefectBatchViewSet.write_off action.
    """

    queryset = DefectWriteOffV2.objects.select_related(
        'batch__product', 'performed_by'
    ).all()
    serializer_class = DefectWriteOffSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['batch', 'reason', 'performed_by']
    ordering_fields = ['performed_at', 'qty_written_off']
    ordering = ['-performed_at']


# ============================================================================
# Analytics and Statistics
# ============================================================================

class DefectAnalyticsViewSet(viewsets.ViewSet):
    """
    ViewSet for defect analytics and statistics.

    Available actions:
    - statistics: Get comprehensive defect statistics
    - export_excel: Export statistics to Excel
    - export_pdf: Export statistics to PDF
    """

    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """
        Get comprehensive defect statistics.

        Query parameters:
        - start_date: Filter from this date (ISO format)
        - end_date: Filter to this date (ISO format)
        - product_id: Filter by product
        - warehouse_name: Filter by warehouse name
        - status: Filter by batch status
        """
        # Parse filters
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        product_id = request.query_params.get('product_id')
        warehouse_id = request.query_params.get('warehouse_name')  # Changed from warehouse_id
        batch_status = request.query_params.get('status')

        if start_date:
            try:
                start_date = datetime.fromisoformat(start_date)
            except ValueError:
                start_date = None

        if end_date:
            try:
                end_date = datetime.fromisoformat(end_date)
            except ValueError:
                end_date = None

        # Get statistics
        stats = DefectAnalyticsService.get_statistics(
            start_date=start_date,
            end_date=end_date,
            product_id=product_id,
            warehouse_id=warehouse_id,
            status=batch_status
        )

        serializer = DefectStatisticsSerializer(stats)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='export/excel')
    def export_excel(self, request):
        """
        Export defect statistics to Excel.

        Uses same filters as statistics endpoint.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse

        # Get statistics
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        product_id = request.query_params.get('product_id')
        warehouse_id = request.query_params.get('warehouse_name')  # Changed from warehouse_id
        batch_status = request.query_params.get('status')

        if start_date:
            try:
                start_date = datetime.fromisoformat(start_date)
            except ValueError:
                start_date = None

        if end_date:
            try:
                end_date = datetime.fromisoformat(end_date)
            except ValueError:
                end_date = None

        stats = DefectAnalyticsService.get_statistics(
            start_date=start_date,
            end_date=end_date,
            product_id=product_id,
            warehouse_id=warehouse_id,
            status=batch_status
        )

        # Create workbook
        wb = Workbook()

        # Sheet 1: Summary
        ws1 = wb.active
        ws1.title = 'Summary'

        # Header style
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        # Summary data
        ws1['A1'] = 'Defect Statistics Summary'
        ws1['A1'].font = Font(size=14, bold=True)

        ws1['A3'] = 'Metric'
        ws1['B3'] = 'Value'
        ws1['A3'].fill = header_fill
        ws1['B3'].fill = header_fill
        ws1['A3'].font = header_font
        ws1['B3'].font = header_font

        totals = stats['totals']
        ws1.append(['Total Batches', totals['total_batches']])
        ws1.append(['Total Defect Quantity', totals['total_qty']])
        ws1.append(['Total Repairable', totals['total_repairable']])
        ws1.append(['Total Non-Repairable', totals['total_non_repairable']])
        ws1.append(['Total Repaired', totals['total_repaired']])
        ws1.append(['Total Written Off', totals['total_written_off']])
        ws1.append(['Remaining Unprocessed', totals['remaining_qty']])
        ws1.append(['Repair Cost (USD)', f"${totals['total_repair_cost_usd']:.2f}"])
        ws1.append(['Recovery Revenue (USD)', f"${totals['total_recovery_revenue_usd']:.2f}"])

        # Sheet 2: By Product
        ws2 = wb.create_sheet('By Product')
        ws2.append(['Product', 'SKU', 'Total Batches', 'Total Qty', 'Repairable', 'Non-Repairable'])
        for row in ws2[1]:
            row.fill = header_fill
            row.font = header_font

        for item in stats['by_product']:
            ws2.append([
                item['product_name'],
                item['product_sku'],
                item['batch_count'],
                item['total_qty'],
                item['repairable_qty'],
                item['non_repairable_qty']
            ])

        # Sheet 3: By Defect Type
        ws3 = wb.create_sheet('By Defect Type')
        ws3.append(['Defect Type', 'Category', 'Total Quantity', 'Affected Batches'])
        for row in ws3[1]:
            row.fill = header_fill
            row.font = header_font

        for item in stats['by_defect_type']:
            ws3.append([
                item['defect_type_name'],
                item['category'],
                item['total_qty'],
                item['batch_count']
            ])

        # Sheet 4: By Status
        ws4 = wb.create_sheet('By Status')
        ws4.append(['Status', 'Count', 'Total Qty'])
        for row in ws4[1]:
            row.fill = header_fill
            row.font = header_font

        for item in stats['by_status']:
            ws4.append([
                item['status'],
                item['count'],
                item['qty_sum']
            ])

        # Sheet 5: Spare Parts Consumption
        ws5 = wb.create_sheet('Spare Parts')
        ws5.append(['Spare Part', 'Unit', 'Total Used', 'Total Cost (USD)'])
        for row in ws5[1]:
            row.fill = header_fill
            row.font = header_font

        for item in stats['spare_parts_consumption']:
            ws5.append([
                item['spare_part_name'],
                item['unit'],
                item['total_qty_used'],
                f"${item['total_cost_usd']:.2f}" if item['total_cost_usd'] else 'N/A'
            ])

        # Auto-size columns for all sheets
        for ws in [ws1, ws2, ws3, ws4, ws5]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'defect_statistics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'

        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='export/pdf')
    def export_pdf(self, request):
        """
        Export defect statistics to PDF.

        Uses same filters as statistics endpoint.
        """
        # TODO: Implement PDF export using ReportLab or WeasyPrint
        # For now, return placeholder
        return Response(
            {'detail': 'PDF export will be implemented in next phase'},
            status=status.HTTP_501_NOT_IMPLEMENTED
        )
