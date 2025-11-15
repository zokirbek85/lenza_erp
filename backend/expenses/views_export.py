"""
Expenses export module - PDF and XLSX generators for list and monthly summaries.
"""
from io import BytesIO
from decimal import Decimal

from django.http import FileResponse, HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from rest_framework.response import Response
from rest_framework.views import APIView
from weasyprint import HTML

from core.permissions import IsAdmin, IsAccountant, IsOwner
from core.utils.company_info import get_company_info
from .export_helpers import (
    describe_export_filters,
    expense_rows_from_queryset,
    get_filtered_expense_queryset,
    total_amounts,
)
from .report_utils import (
    aggregate_monthly_expenses,
    build_monthly_payload,
    build_pdf_context,
    render_pdf_response,
    resolve_date_range,
)


class ExpenseExportBaseView(APIView):
    permission_classes = [IsAdmin | IsAccountant | IsOwner]
    format_kwarg = None

    def _build_logo(self, request, company):
        logo = company.get('logo')
        if logo and not logo.startswith('http'):
            try:
                logo = request.build_absolute_uri(logo)
            except Exception:
                logo = None
        return logo


class ExpenseListPDFExportView(ExpenseExportBaseView):
    def get(self, request):
        queryset = get_filtered_expense_queryset(request).select_related('type')
        rows = expense_rows_from_queryset(queryset)
        totals = total_amounts(queryset)
        filters = describe_export_filters(request)
        company = get_company_info()

        context = {
            'company': company,
            'company_logo': self._build_logo(request, company),
            'generated_at': timezone.now(),
            'rows': rows,
            'filters': filters,
            'totals': totals,
        }
        html = render_to_string('expenses/export_list_pdf.html', context)
        pdf_bytes = HTML(string=html).write_pdf()
        filename = f"expenses_export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"

        return FileResponse(BytesIO(pdf_bytes), as_attachment=True, filename=filename)


class ExpenseListExcelExportView(ExpenseExportBaseView):
    def get(self, request):
        queryset = get_filtered_expense_queryset(request).select_related('type')
        rows = expense_rows_from_queryset(queryset)
        totals = total_amounts(queryset)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Expenses'
        small_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        headers = ['Sana', 'Turi', 'USD', 'UZS', 'Usul', 'Holat', 'Izoh']
        header_font = Font(bold=True)

        worksheet.append(headers)
        for col in range(1, len(headers) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = small_border

        for row_index, row in enumerate(rows, start=2):
            worksheet.cell(row=row_index, column=1, value=row['date'].strftime('%Y-%m-%d')).border = small_border
            worksheet.cell(row=row_index, column=2, value=row['category']).border = small_border
            worksheet.cell(row=row_index, column=3, value=row['amount_usd']).border = small_border
            worksheet.cell(row=row_index, column=4, value=row['amount_uzs']).border = small_border
            worksheet.cell(row=row_index, column=5, value=row['method']).border = small_border
            worksheet.cell(row=row_index, column=6, value=row['status']).border = small_border
            worksheet.cell(row=row_index, column=7, value=row['description'][:100]).border = small_border

        total_row = len(rows) + 2
        worksheet.cell(row=total_row, column=1, value='Jami USD').font = header_font
        worksheet.cell(row=total_row, column=3, value=totals['usd']).border = small_border
        worksheet.cell(row=total_row + 1, column=1, value='Jami UZS').font = header_font
        worksheet.cell(row=total_row + 1, column=4, value=totals['uzs']).border = small_border

        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = max(15, min(max_length + 2, 40))

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        filename = f"expenses_export_{timezone.now().strftime('%Y%m%d')}.xlsx"

        response = HttpResponse(
            stream.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class MonthlyExpensePDFExportView(ExpenseExportBaseView):
    def get(self, request):
        start_date, end_date, month_label = resolve_date_range(
            request.query_params.get('month'),
            request.query_params.get('from'),
            request.query_params.get('to'),
        )
        payload = build_monthly_payload(start_date, end_date, month_label)
        context = build_pdf_context(request, payload)
        return render_pdf_response(
            request,
            context,
            month_label,
            template='expenses/export_monthly_pdf.html',
        )


class MonthlyExpenseExcelExportView(ExpenseExportBaseView):
    def get(self, request):
        start_date, end_date, month_label = resolve_date_range(
            request.query_params.get('month'),
            request.query_params.get('from'),
            request.query_params.get('to'),
        )
        payload = build_monthly_payload(start_date, end_date, month_label)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Monthly Expenses'
        headers = ['Kategoriya', 'USD', 'UZS', 'Ulush (%)']
        worksheet.append(headers)

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col_index in range(1, len(headers) + 1):
            cell = worksheet.cell(row=1, column=col_index)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for idx, row in enumerate(payload['rows'], start=2):
            worksheet.cell(row=idx, column=1, value=row['type']).border = border
            worksheet.cell(row=idx, column=2, value=row['usd']).border = border
            worksheet.cell(row=idx, column=3, value=row['uzs']).border = border
            worksheet.cell(row=idx, column=4, value=row['percentage']).border = border

        total_row = len(payload['rows']) + 2
        worksheet.cell(row=total_row, column=1, value='Jami USD').font = Font(bold=True)
        worksheet.cell(row=total_row, column=2, value=payload['total_usd']).border = border
        worksheet.cell(row=total_row + 1, column=1, value='Jami UZS').font = Font(bold=True)
        worksheet.cell(row=total_row + 1, column=2, value=payload['total_uzs']).border = border
        worksheet.cell(row=total_row + 2, column=1, value='Kurs (1 USD)').font = Font(bold=True)
        worksheet.cell(row=total_row + 2, column=2, value=payload.get('usd_rate', 0)).border = border

        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = max(15, min(max_length + 2, 40))

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        filename = f"monthly_expenses_{month_label}.xlsx"

        response = HttpResponse(
            stream.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class LegacyExpenseExportView(ExpenseExportBaseView):
    def get(self, request):
        return Response(
            {
                'detail': 'Use /api/expenses/export/pdf/ or /api/expenses/export/excel/ for data exports.',
            },
            status=200,
        )
