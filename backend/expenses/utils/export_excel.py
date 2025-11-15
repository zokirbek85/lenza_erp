"""
Expenses Export Utilities - Excel Generator
"""
from datetime import datetime
from decimal import Decimal
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from ..models import Expense
from core.models import CompanyInfo
from payments.models import CurrencyRate


def generate_expense_excel(date_from=None, date_to=None, currency='USD'):
    """
    Excel eksport - chiqimlar ro'yxati
    """
    # Ma'lumotlarni olish
    qs = Expense.objects.filter(status=Expense.STATUS_APPROVED).select_related('type', 'card', 'created_by')
    
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if currency:
        qs = qs.filter(currency=currency)
    
    expenses = qs.order_by('date')
    
    # Workbook yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = "Chiqimlar"
    
    # Kompaniya ma'lumotlari
    company = CompanyInfo.objects.first()
    
    # Header styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Kompaniya nomi
    if company:
        ws.merge_cells('A1:G1')
        cell = ws['A1']
        cell.value = company.firm_name
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center')
    
    # Sarlavha
    ws.merge_cells('A2:G2')
    cell = ws['A2']
    cell.value = f"CHIQIMLAR RO'YXATI"
    cell.font = Font(bold=True, size=13)
    cell.alignment = Alignment(horizontal='center')
    
    # Davr
    period_text = f"Davr: {date_from or 'Boshidan'} - {date_to or 'Hozirgacha'}"
    ws.merge_cells('A3:G3')
    cell = ws['A3']
    cell.value = period_text
    cell.alignment = Alignment(horizontal='center')
    
    # Bo'sh qator
    current_row = 5
    
    # Table headers
    headers = ['№', 'Sana', 'Turi', 'To\'lov usuli', 'Karta', 'Summa', 'Tavsif']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    current_row += 1
    
    # Ma'lumotlar
    total = Decimal('0.00')
    for idx, expense in enumerate(expenses, 1):
        ws.cell(row=current_row, column=1, value=idx).border = border
        ws.cell(row=current_row, column=2, value=expense.date.strftime('%d.%m.%Y')).border = border
        ws.cell(row=current_row, column=3, value=expense.type.name).border = border
        ws.cell(row=current_row, column=4, value=expense.get_method_display()).border = border
        ws.cell(row=current_row, column=5, value=expense.card.name if expense.card else '-').border = border
        ws.cell(row=current_row, column=6, value=float(expense.amount)).border = border
        ws.cell(row=current_row, column=7, value=expense.description or '-').border = border
        
        total += expense.amount
        current_row += 1
    
    # Jami
    ws.merge_cells(f'A{current_row}:E{current_row}')
    cell = ws.cell(row=current_row, column=1)
    cell.value = 'JAMI:'
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='right')
    cell.border = border
    
    cell = ws.cell(row=current_row, column=6)
    cell.value = float(total)
    cell.font = Font(bold=True)
    cell.border = border
    
    cell = ws.cell(row=current_row, column=7)
    cell.value = currency
    cell.font = Font(bold=True)
    cell.border = border
    
    # USD konvertatsiya (agar UZS bo'lsa)
    if currency == 'UZS':
        current_row += 1
        rate = CurrencyRate.objects.order_by('-rate_date').first()
        usd_rate = Decimal(str(rate.usd_to_uzs)) if rate else Decimal('12500')
        total_usd = (total / usd_rate).quantize(Decimal('0.01'))
        
        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws.cell(row=current_row, column=1)
        cell.value = 'USD kursi bo\'yicha:'
        cell.alignment = Alignment(horizontal='right')
        
        cell = ws.cell(row=current_row, column=6)
        cell.value = float(total_usd)
        
        cell = ws.cell(row=current_row, column=7)
        cell.value = 'USD'
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 30
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"chiqimlar_{date_from or 'all'}_{date_to or 'all'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
