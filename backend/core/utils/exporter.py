from io import BytesIO

from openpyxl import Workbook

from .excel_export import prepare_workbook, workbook_to_bytes
from .temp_files import save_temp_file


def _prepare_workbook(title: str, headers: list[str]):
    """Legacy wrapper - use prepare_workbook from excel_export instead."""
    return prepare_workbook(title, headers)


def export_orders_to_excel(orders):
    """
    Export orders with detailed information including order items.
    Creates a comprehensive report with multiple sheets.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    workbook = Workbook()
    
    # Summary sheet
    ws_summary = workbook.active
    ws_summary.title = 'Buyurtmalar'
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Headers for summary sheet
    summary_headers = [
        'Buyurtma №',
        'Diller',
        'Status',
        'Mahsulotlar soni',
        'Jami summa (USD)',
        'Sana',
        'Yaratgan',
        'Yaratilgan vaqti'
    ]
    
    ws_summary.append(summary_headers)
    
    # Style header row
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add summary data
    for order in orders:
        items_count = order.items.count() if hasattr(order, 'items') else 0
        ws_summary.append([
            getattr(order, 'display_no', ''),
            getattr(order.dealer, 'name', '') if order.dealer else '',
            order.get_status_display() if hasattr(order, 'get_status_display') else order.status,
            items_count,
            float(order.total_usd),
            order.value_date.strftime('%Y-%m-%d') if order.value_date else '',
            getattr(order.created_by, 'username', '') if order.created_by else '',
            order.created_at.strftime('%Y-%m-%d %H:%M') if order.created_at else '',
        ])
    
    # Auto-adjust column widths
    for column in ws_summary.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # Detailed items sheet
    ws_details = workbook.create_sheet('Mahsulotlar detali')
    
    detail_headers = [
        'Buyurtma №',
        'Diller',
        'Mahsulot',
        'SKU',
        'Miqdor',
        'Narx (USD)',
        'Chegirma (%)',
        'Jami (USD)',
        'Sana'
    ]
    
    ws_details.append(detail_headers)
    
    # Style header row
    for cell in ws_details[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add detailed items data
    for order in orders:
        if hasattr(order, 'items'):
            for item in order.items.all():
                product_name = getattr(item.product, 'name', '') if item.product else ''
                product_sku = getattr(item.product, 'sku', '') if item.product else ''
                
                # Calculate item total with discount
                item_total = float(item.price_usd * item.qty)
                if hasattr(item, 'discount_percent') and item.discount_percent:
                    discount_amount = item_total * (float(item.discount_percent) / 100)
                    item_total -= discount_amount
                
                ws_details.append([
                    getattr(order, 'display_no', ''),
                    getattr(order.dealer, 'name', '') if order.dealer else '',
                    product_name,
                    product_sku,
                    float(item.qty),
                    float(item.price_usd),
                    float(getattr(item, 'discount_percent', 0)),
                    item_total,
                    order.value_date.strftime('%Y-%m-%d') if order.value_date else '',
                ])
    
    # Auto-adjust column widths for details sheet
    for column in ws_details.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_details.column_dimensions[column_letter].width = adjusted_width
    
    return _workbook_to_file(workbook, 'orders')


def export_products_to_excel(products):
    workbook, worksheet = _prepare_workbook(
        'Products',
        ['SKU', 'Name', 'Brand', 'Category', 'Sell USD', 'Stock OK', 'Stock Defect'],
    )
    for product in products:
        worksheet.append(
            [
                product.sku,
                product.name,
                getattr(product.brand, 'name', ''),
                getattr(product.category, 'name', ''),
                float(product.sell_price_usd),
                float(product.stock_ok or 0),
                float(product.stock_defect or 0),
            ]
        )
    return _workbook_to_file(workbook, 'products')


# export_payments_to_excel removed - Payment module deleted

def export_returns_to_excel(returns):
    workbook, worksheet = _prepare_workbook(
        'Returns',
        ['Dealer', 'Product', 'Quantity', 'Return Type', 'Reason', 'Created At'],
    )
    for entry in returns:
        worksheet.append(
            [
                getattr(entry.dealer, 'name', ''),
                getattr(entry.product, 'name', ''),
                float(entry.quantity or 0),
                entry.get_return_type_display() if hasattr(entry, 'get_return_type_display') else entry.return_type,
                entry.reason or '',
                entry.created_at.isoformat() if entry.created_at else '',
            ]
        )
    return _workbook_to_file(workbook, 'returns')


# export_expenses_to_excel removed - Expense module deleted

def _workbook_to_file(workbook: Workbook, prefix: str):
    """
    Save openpyxl Workbook to temporary file.
    
    CRITICAL: Do NOT add UTF-8 BOM to XLSX files!
    XLSX is a ZIP archive and must start with PK signature (0x504B0304).
    Adding BOM corrupts the ZIP structure and Excel cannot open the file.
    
    UTF-8 BOM is only for CSV/TXT files, never for XLSX.
    """
    content = workbook_to_bytes(workbook)
    return save_temp_file(content, prefix, '.xlsx')


def export_reconciliation_to_excel(data: dict, detailed: bool = False, language: str = 'uz'):
    """
    Build an Excel workbook from reconciliation data structure returned by
    services.reconciliation.get_reconciliation_data.
    When detailed=True, include per-order item lines on separate sheets or indented rows.
    language: 'uz', 'ru', or 'en' for internationalization.
    """
    # Translation dictionaries
    translations = {
        'uz': {
            'summary': 'Xulosa',
            'orders': 'Buyurtmalar',
            'payments': 'To\'lovlar',
            'returns': 'Qaytarishlar',
            'order_items': 'Buyurtma detallari',
            'dealer': 'Diler',
            'code': 'Kod',
            'period': 'Davr',
            'opening_balance': 'Boshlang\'ich balans (USD)',
            'closing_balance': 'Yakuniy balans (USD)',
            'date': 'Sana',
            'order_no': 'Buyurtma raqami',
            'amount_usd': 'Summa (USD)',
            'method': 'Usul',
            'product': 'Mahsulot',
            'quantity': 'Miqdor',
            'price_usd': 'Narx (USD)',
            'line_total_usd': 'Jami (USD)',
        },
        'ru': {
            'summary': 'Сводка',
            'orders': 'Заказы',
            'payments': 'Платежи',
            'returns': 'Возвраты',
            'order_items': 'Детали заказа',
            'dealer': 'Дилер',
            'code': 'Код',
            'period': 'Период',
            'opening_balance': 'Начальный баланс (USD)',
            'closing_balance': 'Конечный баланс (USD)',
            'date': 'Дата',
            'order_no': 'Номер заказа',
            'amount_usd': 'Сумма (USD)',
            'method': 'Метод',
            'product': 'Продукт',
            'quantity': 'Количество',
            'price_usd': 'Цена (USD)',
            'line_total_usd': 'Итого (USD)',
        },
        'en': {
            'summary': 'Summary',
            'orders': 'Orders',
            'payments': 'Payments',
            'returns': 'Returns',
            'order_items': 'Order Items',
            'dealer': 'Dealer',
            'code': 'Code',
            'period': 'Period',
            'opening_balance': 'Opening Balance (USD)',
            'closing_balance': 'Closing Balance (USD)',
            'date': 'Date',
            'order_no': 'Order #',
            'amount_usd': 'Amount (USD)',
            'method': 'Method',
            'product': 'Product',
            'quantity': 'Quantity',
            'price_usd': 'Price (USD)',
            'line_total_usd': 'Line Total (USD)',
        },
    }
    
    # Get translations for selected language
    t = translations.get(language, translations['uz'])
    
    workbook = Workbook()
    ws_summary = workbook.active
    ws_summary.title = t['summary']

    # Header info
    ws_summary.append([t['dealer'], data.get('dealer', '')])
    ws_summary.append([t['code'], data.get('dealer_code', '')])
    ws_summary.append([t['period'], data.get('period', '')])
    ws_summary.append([])

    # Balances
    ws_summary.append([t['opening_balance'], float(data.get('opening_balance', 0))])
    ws_summary.append([t['closing_balance'], float(data.get('closing_balance', 0))])
    ws_summary.append([])

    # Orders sheet
    ws_orders = workbook.create_sheet(t['orders'])
    ws_orders.append([t['date'], t['order_no'], t['amount_usd']])
    for row in data.get('orders', []) or []:
        ws_orders.append([
            row.get('date'),
            row.get('order_no'),
            float(row.get('amount_usd', 0)),
        ])

    # Payments sheet
    ws_payments = workbook.create_sheet(t['payments'])
    ws_payments.append([t['date'], t['method'], t['amount_usd']])
    for row in data.get('payments', []) or []:
        ws_payments.append([
            row.get('date'),
            row.get('method'),
            float(row.get('amount_usd', 0)),
        ])

    # Returns sheet
    ws_returns = workbook.create_sheet(t['returns'])
    ws_returns.append([t['date'], t['order_no'], t['amount_usd']])
    for row in data.get('returns', []) or []:
        ws_returns.append([
            row.get('date'),
            row.get('order_no'),
            float(row.get('amount_usd', 0)),
        ])

    if detailed and data.get('orders_detailed'):
        ws_details = workbook.create_sheet(t['order_items'])
        ws_details.append([t['order_no'], t['date'], t['product'], t['quantity'], t['price_usd'], t['line_total_usd']])
        for od in data.get('orders_detailed', []) or []:
            order_no = od.get('order_number')
            order_date = od.get('date')
            for item in od.get('items', []) or []:
                ws_details.append([
                    order_no,
                    order_date,
                    item.get('product_name'),
                    float(item.get('quantity', 0)),
                    float(item.get('price', 0)),
                    float(item.get('total', 0)),
                ])

    return _workbook_to_file(workbook, 'reconciliation')