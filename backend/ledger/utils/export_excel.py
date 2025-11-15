"""
Ledger Export Utilities - Excel Generator
"""
from datetime import datetime
from decimal import Decimal
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from ..models import LedgerRecord
from core.models import CompanyInfo
from payments.models import CurrencyRate


def generate_ledger_excel(date_from=None, date_to=None, currency='USD'):
    """
    Excel eksport - kassa operatsiyalari
    """
    # Ma'lumotlarni olish
    qs = LedgerRecord.objects.select_related('created_by').all()
    
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if currency:
        qs = qs.filter(currency=currency)
    
    records = qs.order_by('date')
    
    # Workbook yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = "Kassa"
    
    # Kompaniya ma'lumotlari
    company = CompanyInfo.objects.first()
    
    # Header styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    income_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    expense_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Kompaniya nomi
    if company:
        ws.merge_cells('A1:G1')
        cell = ws['A1']
        cell.value = company.firm_name
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center')
    
    # Sarlavha
    ws.merge_cells('A2:G2')
    cell = ws['A2']
    cell.value = f"KASSA OPERATSIYALARI"
    cell.font = Font(bold=True, size=13)
    cell.alignment = Alignment(horizontal='center')
    
    # Davr
    period_text = f"Davr: {date_from or 'Boshidan'} - {date_to or 'Hozirgacha'}"
    ws.merge_cells('A3:G3')
    cell = ws['A3']
    cell.value = period_text
    cell.alignment = Alignment(horizontal='center')
    
    # Bo'sh qator
    current_row = 5
    
    # Table headers
    headers = ['№', 'Sana', 'Turi', 'Manba', 'Summa', 'Balans', 'Tavsif']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    current_row += 1
    
    # Ma'lumotlar
    cumulative_balance = Decimal('0.00')
    total_income = Decimal('0.00')
    total_expense = Decimal('0.00')
    
    for idx, record in enumerate(records, 1):
        # Balans hisoblash
        if record.type == LedgerRecord.TYPE_INCOME:
            cumulative_balance += record.amount
            total_income += record.amount
            row_fill = income_fill
        else:
            cumulative_balance -= record.amount
            total_expense += record.amount
            row_fill = expense_fill
        
        ws.cell(row=current_row, column=1, value=idx).border = border
        ws.cell(row=current_row, column=2, value=record.date.strftime('%d.%m.%Y')).border = border
        
        cell = ws.cell(row=current_row, column=3, value=record.get_type_display())
        cell.border = border
        cell.fill = row_fill
        
        ws.cell(row=current_row, column=4, value=record.get_source_display()).border = border
        ws.cell(row=current_row, column=5, value=float(record.amount)).border = border
        ws.cell(row=current_row, column=6, value=float(cumulative_balance)).border = border
        ws.cell(row=current_row, column=7, value=record.description or '-').border = border
        
        current_row += 1
    
    # Bo'sh qator
    current_row += 1
    
    # Jami - Kirimlar
    ws.merge_cells(f'A{current_row}:D{current_row}')
    cell = ws.cell(row=current_row, column=1)
    cell.value = 'JAMI KIRIMLAR:'
    cell.font = Font(bold=True, color="155724")
    cell.alignment = Alignment(horizontal='right')
    cell.border = border
    cell.fill = income_fill
    
    cell = ws.cell(row=current_row, column=5)
    cell.value = float(total_income)
    cell.font = Font(bold=True, color="155724")
    cell.border = border
    cell.fill = income_fill
    
    ws.cell(row=current_row, column=6, value=currency).border = border
    ws.cell(row=current_row, column=7).border = border
    
    current_row += 1
    
    # Jami - Chiqimlar
    ws.merge_cells(f'A{current_row}:D{current_row}')
    cell = ws.cell(row=current_row, column=1)
    cell.value = 'JAMI CHIQIMLAR:'
    cell.font = Font(bold=True, color="721C24")
    cell.alignment = Alignment(horizontal='right')
    cell.border = border
    cell.fill = expense_fill
    
    cell = ws.cell(row=current_row, column=5)
    cell.value = float(total_expense)
    cell.font = Font(bold=True, color="721C24")
    cell.border = border
    cell.fill = expense_fill
    
    ws.cell(row=current_row, column=6, value=currency).border = border
    ws.cell(row=current_row, column=7).border = border
    
    current_row += 1
    
    # Balans
    balance = total_income - total_expense
    ws.merge_cells(f'A{current_row}:D{current_row}')
    cell = ws.cell(row=current_row, column=1)
    cell.value = 'BALANS:'
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='right')
    cell.border = border
    
    cell = ws.cell(row=current_row, column=5)
    cell.value = float(balance)
    cell.font = Font(bold=True, size=12)
    cell.border = border
    
    cell = ws.cell(row=current_row, column=6)
    cell.value = currency
    cell.font = Font(bold=True)
    cell.border = border
    
    ws.cell(row=current_row, column=7).border = border
    
    # USD konvertatsiya (agar UZS bo'lsa)
    if currency == 'UZS':
        current_row += 1
        rate = CurrencyRate.objects.order_by('-rate_date').first()
        usd_rate = Decimal(str(rate.usd_to_uzs)) if rate else Decimal('12500')
        balance_usd = (balance / usd_rate).quantize(Decimal('0.01'))
        
        ws.merge_cells(f'A{current_row}:D{current_row}')
        cell = ws.cell(row=current_row, column=1)
        cell.value = 'USD kursi bo\'yicha:'
        cell.alignment = Alignment(horizontal='right')
        
        cell = ws.cell(row=current_row, column=5)
        cell.value = float(balance_usd)
        
        cell = ws.cell(row=current_row, column=6)
        cell.value = 'USD'
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 35
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"kassa_{date_from or 'all'}_{date_to or 'all'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
