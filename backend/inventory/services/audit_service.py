"""
Inventory Audit Service
Handles export/import of physical inventory count data and stock adjustments.
"""
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any, Dict, List

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from catalog.models import Product
from inventory.models import InventoryAdjustment


class AuditExportService:
    """
    Service for exporting current inventory state to Excel for physical audit.
    """
    
    # Excel column headers
    HEADERS = [
        'SKU',
        'Name',
        'System Stock OK',
        'System Stock Defect',
        'Real Stock OK',
        'Real Stock Defect',
    ]
    
    # Column widths for better readability
    COLUMN_WIDTHS = {
        'A': 15,  # SKU
        'B': 40,  # Name
        'C': 18,  # System Stock OK
        'D': 20,  # System Stock Defect
        'E': 18,  # Real Stock OK
        'F': 20,  # Real Stock Defect
    }
    
    @classmethod
    def export_to_excel(cls) -> BytesIO:
        """
        Export all products with current stock levels to Excel.
        
        Returns:
            BytesIO: Excel file buffer
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Audit"
        
        # Write headers
        for col_idx, header in enumerate(cls.HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        for col_letter, width in cls.COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width
        
        # Fetch all active products ordered by SKU
        products = Product.objects.filter(is_active=True).order_by('sku')
        
        # Write product data
        for row_idx, product in enumerate(products, start=2):
            ws.cell(row=row_idx, column=1, value=product.sku)
            ws.cell(row=row_idx, column=2, value=product.name)
            ws.cell(row=row_idx, column=3, value=int(product.stock_ok))
            ws.cell(row=row_idx, column=4, value=int(product.stock_defect))
            # Columns E and F (Real Stock) are left empty for manual entry
            ws.cell(row=row_idx, column=5, value=None)
            ws.cell(row=row_idx, column=6, value=None)
            
            # Style data cells
            for col_idx in range(1, 7):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal='left' if col_idx <= 2 else 'center', vertical='center')
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer


class AuditImportService:
    """
    Service for importing physical audit results and creating stock adjustments.
    """
    
    # Expected column indices (0-based)
    COL_SKU = 0
    COL_NAME = 1
    COL_SYSTEM_OK = 2
    COL_SYSTEM_DEFECT = 3
    COL_REAL_OK = 4
    COL_REAL_DEFECT = 5
    
    MIN_COLUMNS = 6
    
    @classmethod
    def validate_file_format(cls, file_data: BytesIO) -> Dict[str, Any]:
        """
        Validate Excel file structure.
        
        Args:
            file_data: Excel file buffer
            
        Returns:
            Dict with 'valid' boolean and 'error' message if invalid
        """
        try:
            wb = load_workbook(file_data, read_only=True, data_only=True)
            ws = wb.active
            
            # Check if file has data
            if ws.max_row < 2:
                return {'valid': False, 'error': 'File is empty or has no data rows'}
            
            # Check column count
            if ws.max_column < cls.MIN_COLUMNS:
                return {'valid': False, 'error': f'File must have at least {cls.MIN_COLUMNS} columns'}
            
            # Validate headers (first row)
            headers = [cell.value for cell in ws[1]]
            expected_headers = AuditExportService.HEADERS
            
            for idx, expected in enumerate(expected_headers):
                if idx >= len(headers) or headers[idx] != expected:
                    return {'valid': False, 'error': f'Invalid header at column {idx + 1}. Expected "{expected}", got "{headers[idx] if idx < len(headers) else "None"}"'}
            
            return {'valid': True}
            
        except Exception as e:
            return {'valid': False, 'error': f'Failed to read file: {str(e)}'}
    
    @classmethod
    def process_audit_import(
        cls,
        file_data: BytesIO,
        user,
        audit_date: date = None,
        comment: str = ''
    ) -> Dict[str, Any]:
        """
        Process audit import file and create stock adjustments.
        
        Args:
            file_data: Excel file buffer
            user: User performing the audit
            audit_date: Date of physical audit (defaults to today)
            comment: Optional comment for all adjustments
            
        Returns:
            Dict with statistics:
                - total_products: Total products in file
                - updated_products: Products with changes
                - unchanged_products: Products with no changes
                - adjustments: List of adjustment details
                - errors: List of error messages
        """
        if audit_date is None:
            audit_date = date.today()
        
        # Validate file format first
        file_data.seek(0)
        validation = cls.validate_file_format(file_data)
        if not validation['valid']:
            return {
                'success': False,
                'error': validation['error'],
            }
        
        # Reset buffer and load for processing
        file_data.seek(0)
        wb = load_workbook(file_data, data_only=True)
        ws = wb.active
        
        results = {
            'success': True,
            'total_products': 0,
            'updated_products': 0,
            'unchanged_products': 0,
            'adjustments': [],
            'errors': [],
        }
        
        # Process each row (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < cls.MIN_COLUMNS:
                continue
            
            sku = row[cls.COL_SKU]
            real_ok = row[cls.COL_REAL_OK]
            real_defect = row[cls.COL_REAL_DEFECT]
            
            # Skip if Real Stock columns are empty
            if real_ok is None and real_defect is None:
                continue
            
            # Validate SKU
            if not sku:
                results['errors'].append(f'Row {row_idx}: SKU is empty')
                continue
            
            # Validate Real Stock values
            try:
                real_ok = Decimal(str(real_ok)) if real_ok is not None else None
                real_defect = Decimal(str(real_defect)) if real_defect is not None else None
            except (ValueError, TypeError, Exception):
                results['errors'].append(f'Row {row_idx}: Real Stock values must be numbers')
                continue
            
            # Skip if both values are None after conversion
            if real_ok is None and real_defect is None:
                continue
            
            results['total_products'] += 1
            
            # Process single product in atomic transaction
            try:
                adjustment = cls._process_single_product(
                    sku=sku,
                    real_ok=real_ok,
                    real_defect=real_defect,
                    user=user,
                    audit_date=audit_date,
                    comment=comment,
                )
                
                if adjustment:
                    results['updated_products'] += 1
                    results['adjustments'].append({
                        'sku': sku,
                        'product_name': adjustment.product.name,
                        'delta_ok': adjustment.delta_ok,
                        'delta_defect': adjustment.delta_defect,
                        'previous_ok': adjustment.previous_ok,
                        'previous_defect': adjustment.previous_defect,
                        'new_ok': adjustment.new_ok,
                        'new_defect': adjustment.new_defect,
                    })
                else:
                    results['unchanged_products'] += 1
                    
            except Product.DoesNotExist:
                results['errors'].append(f'Row {row_idx}: Product with SKU "{sku}" not found')
            except Exception as e:
                results['errors'].append(f'Row {row_idx}: {str(e)}')
        
        return results
    
    @classmethod
    @transaction.atomic
    def _process_single_product(
        cls,
        sku: str,
        real_ok: Decimal | None,
        real_defect: Decimal | None,
        user,
        audit_date: date,
        comment: str,
    ) -> InventoryAdjustment | None:
        """
        Process single product adjustment within atomic transaction.
        
        Args:
            sku: Product SKU
            real_ok: Real stock OK from physical count (Decimal)
            real_defect: Real stock defect from physical count (Decimal)
            user: User performing audit
            audit_date: Audit date
            comment: Optional comment
            
        Returns:
            InventoryAdjustment object if created, None if no changes
            
        Raises:
            Product.DoesNotExist: If product not found
        """
        # Lock product row for update
        product = Product.objects.select_for_update().get(sku=sku)
        
        # Get current stock values as Decimal
        previous_ok = product.stock_ok
        previous_defect = product.stock_defect
        
        # Use previous values if real values not provided
        new_ok = real_ok if real_ok is not None else previous_ok
        new_defect = real_defect if real_defect is not None else previous_defect
        
        # Calculate deltas
        delta_ok = new_ok - previous_ok
        delta_defect = new_defect - previous_defect
        
        # Skip if no changes
        if delta_ok == 0 and delta_defect == 0:
            return None
        
        # Update product stock with real values
        product.stock_ok = new_ok
        product.stock_defect = new_defect
        product.save(update_fields=['stock_ok', 'stock_defect', 'updated_at'])
        
        # Create adjustment record
        adjustment = InventoryAdjustment.objects.create(
            product=product,
            delta_ok=delta_ok,
            delta_defect=delta_defect,
            previous_ok=previous_ok,
            previous_defect=previous_defect,
            new_ok=new_ok,
            new_defect=new_defect,
            date=audit_date,
            created_by=user,
            comment=comment,
        )
        
        return adjustment
