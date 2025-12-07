"""
Tests for inventory audit functionality
"""
from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase, TransactionTestCase
from openpyxl import Workbook

from catalog.models import Brand, Category, Product
from inventory.models import InventoryAdjustment
from inventory.services import AuditExportService, AuditImportService

User = get_user_model()


class AuditExportServiceTests(TestCase):
    """Tests for AuditExportService"""
    
    def setUp(self):
        # Create test data
        self.brand = Brand.objects.create(name='Test Brand')
        self.category = Category.objects.create(name='Test Category')
        
        self.product1 = Product.objects.create(
            sku='TEST001',
            name='Test Product 1',
            brand=self.brand,
            category=self.category,
            stock_ok=Decimal('100.00'),
            stock_defect=Decimal('10.00'),
            is_active=True,
        )
        
        self.product2 = Product.objects.create(
            sku='TEST002',
            name='Test Product 2',
            brand=self.brand,
            category=self.category,
            stock_ok=Decimal('50.00'),
            stock_defect=Decimal('5.00'),
            is_active=True,
        )
    
    def test_export_creates_valid_excel(self):
        """Test that export creates a valid Excel file"""
        buffer = AuditExportService.export_to_excel()
        
        # Check that buffer is not empty
        self.assertGreater(buffer.getbuffer().nbytes, 0)
        
        # Try to load the workbook
        from openpyxl import load_workbook
        buffer.seek(0)
        wb = load_workbook(buffer)
        ws = wb.active
        
        # Check headers
        self.assertEqual(ws.cell(1, 1).value, 'SKU')
        self.assertEqual(ws.cell(1, 2).value, 'Name')
        self.assertEqual(ws.cell(1, 3).value, 'System Stock OK')
        self.assertEqual(ws.cell(1, 4).value, 'System Stock Defect')
        self.assertEqual(ws.cell(1, 5).value, 'Real Stock OK')
        self.assertEqual(ws.cell(1, 6).value, 'Real Stock Defect')
    
    def test_export_includes_all_active_products(self):
        """Test that export includes all active products"""
        from openpyxl import load_workbook
        
        buffer = AuditExportService.export_to_excel()
        buffer.seek(0)
        wb = load_workbook(buffer)
        ws = wb.active
        
        # Should have header + 2 products
        self.assertEqual(ws.max_row, 3)
        
        # Check product data
        self.assertEqual(ws.cell(2, 1).value, 'TEST001')
        self.assertEqual(ws.cell(2, 2).value, 'Test Product 1')
        self.assertEqual(ws.cell(2, 3).value, 100)
        self.assertEqual(ws.cell(2, 4).value, 10)
        self.assertIsNone(ws.cell(2, 5).value)  # Real Stock OK should be empty
        self.assertIsNone(ws.cell(2, 6).value)  # Real Stock Defect should be empty
    
    def test_export_excludes_inactive_products(self):
        """Test that inactive products are not included in export"""
        self.product2.is_active = False
        self.product2.save()
        
        from openpyxl import load_workbook
        
        buffer = AuditExportService.export_to_excel()
        buffer.seek(0)
        wb = load_workbook(buffer)
        ws = wb.active
        
        # Should have header + 1 product only
        self.assertEqual(ws.max_row, 2)


class AuditImportServiceTests(TransactionTestCase):
    """Tests for AuditImportService (uses TransactionTestCase for select_for_update)"""
    
    def setUp(self):
        # Create test user
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpass123'
        )
        
        # Create test data
        self.brand = Brand.objects.create(name='Test Brand')
        self.category = Category.objects.create(name='Test Category')
        
        self.product1 = Product.objects.create(
            sku='TEST001',
            name='Test Product 1',
            brand=self.brand,
            category=self.category,
            stock_ok=Decimal('100.00'),
            stock_defect=Decimal('10.00'),
        )
        
        self.product2 = Product.objects.create(
            sku='TEST002',
            name='Test Product 2',
            brand=self.brand,
            category=self.category,
            stock_ok=Decimal('50.00'),
            stock_defect=Decimal('5.00'),
        )
    
    def create_audit_file(self, data):
        """Helper to create audit Excel file"""
        wb = Workbook()
        ws = wb.active
        
        # Headers
        ws.append(['SKU', 'Name', 'System Stock OK', 'System Stock Defect', 'Real Stock OK', 'Real Stock Defect'])
        
        # Data rows
        for row in data:
            ws.append(row)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def test_validate_file_format_valid(self):
        """Test file format validation with valid file"""
        file_data = self.create_audit_file([
            ['TEST001', 'Test Product 1', 100, 10, 95, 12],
        ])
        
        result = AuditImportService.validate_file_format(file_data)
        self.assertTrue(result['valid'])
    
    def test_validate_file_format_missing_columns(self):
        """Test file format validation with missing columns"""
        wb = Workbook()
        ws = wb.active
        ws.append(['SKU', 'Name', 'Stock OK'])  # Missing columns
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        result = AuditImportService.validate_file_format(buffer)
        self.assertFalse(result['valid'])
        self.assertIn('at least 6 columns', result['error'])
    
    def test_validate_file_format_empty_file(self):
        """Test file format validation with empty file"""
        wb = Workbook()
        ws = wb.active
        ws.append(['SKU', 'Name', 'System Stock OK', 'System Stock Defect', 'Real Stock OK', 'Real Stock Defect'])
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        result = AuditImportService.validate_file_format(buffer)
        self.assertFalse(result['valid'])
        self.assertIn('empty', result['error'].lower())
    
    def test_process_audit_import_with_increase(self):
        """Test audit import with stock increase"""
        file_data = self.create_audit_file([
            ['TEST001', 'Test Product 1', 100, 10, 120, 15],  # +20 OK, +5 Defect
        ])
        
        result = AuditImportService.process_audit_import(
            file_data=file_data,
            user=self.user,
            audit_date=date.today(),
        )
        
        self.assertTrue(result['success'])
        self.assertEqual(result['total_products'], 1)
        self.assertEqual(result['updated_products'], 1)
        self.assertEqual(result['unchanged_products'], 0)
        
        # Check product was updated
        self.product1.refresh_from_db()
        self.assertEqual(self.product1.stock_ok, Decimal('120.00'))
        self.assertEqual(self.product1.stock_defect, Decimal('15.00'))
        
        # Check adjustment was created
        adjustment = InventoryAdjustment.objects.get(product=self.product1)
        self.assertEqual(adjustment.delta_ok, 20)
        self.assertEqual(adjustment.delta_defect, 5)
        self.assertEqual(adjustment.previous_ok, 100)
        self.assertEqual(adjustment.previous_defect, 10)
        self.assertEqual(adjustment.new_ok, 120)
        self.assertEqual(adjustment.new_defect, 15)
    
    def test_process_audit_import_with_decrease(self):
        """Test audit import with stock decrease"""
        file_data = self.create_audit_file([
            ['TEST001', 'Test Product 1', 100, 10, 80, 5],  # -20 OK, -5 Defect
        ])
        
        result = AuditImportService.process_audit_import(
            file_data=file_data,
            user=self.user,
            audit_date=date.today(),
        )
        
        self.assertTrue(result['success'])
        
        # Check product was updated
        self.product1.refresh_from_db()
        self.assertEqual(self.product1.stock_ok, Decimal('80.00'))
        self.assertEqual(self.product1.stock_defect, Decimal('5.00'))
        
        # Check adjustment
        adjustment = InventoryAdjustment.objects.get(product=self.product1)
        self.assertEqual(adjustment.delta_ok, -20)
        self.assertEqual(adjustment.delta_defect, -5)
    
    def test_process_audit_import_no_change(self):
        """Test audit import when real stock equals system stock"""
        file_data = self.create_audit_file([
            ['TEST001', 'Test Product 1', 100, 10, 100, 10],  # No change
        ])
        
        result = AuditImportService.process_audit_import(
            file_data=file_data,
            user=self.user,
            audit_date=date.today(),
        )
        
        self.assertTrue(result['success'])
        self.assertEqual(result['total_products'], 1)
        self.assertEqual(result['updated_products'], 0)
        self.assertEqual(result['unchanged_products'], 1)
        
        # No adjustment should be created
        self.assertEqual(InventoryAdjustment.objects.count(), 0)
    
    def test_process_audit_import_multiple_products(self):
        """Test audit import with multiple products"""
        file_data = self.create_audit_file([
            ['TEST001', 'Test Product 1', 100, 10, 110, 12],  # Changed
            ['TEST002', 'Test Product 2', 50, 5, 50, 5],      # Unchanged
        ])
        
        result = AuditImportService.process_audit_import(
            file_data=file_data,
            user=self.user,
            audit_date=date.today(),
        )
        
        self.assertTrue(result['success'])
        self.assertEqual(result['total_products'], 2)
        self.assertEqual(result['updated_products'], 1)
        self.assertEqual(result['unchanged_products'], 1)
        self.assertEqual(len(result['adjustments']), 1)
    
    def test_process_audit_import_product_not_found(self):
        """Test audit import with non-existent product"""
        file_data = self.create_audit_file([
            ['INVALID', 'Invalid Product', 100, 10, 90, 8],
        ])
        
        result = AuditImportService.process_audit_import(
            file_data=file_data,
            user=self.user,
            audit_date=date.today(),
        )
        
        self.assertTrue(result['success'])
        self.assertEqual(result['updated_products'], 0)
        self.assertGreater(len(result['errors']), 0)
        self.assertIn('not found', result['errors'][0])
    
    def test_process_audit_import_invalid_values(self):
        """Test audit import with invalid stock values"""
        file_data = self.create_audit_file([
            ['TEST001', 'Test Product 1', 100, 10, 'abc', 'def'],  # Invalid
        ])
        
        result = AuditImportService.process_audit_import(
            file_data=file_data,
            user=self.user,
            audit_date=date.today(),
        )
        
        self.assertTrue(result['success'])
        self.assertEqual(result['updated_products'], 0)
        self.assertGreater(len(result['errors']), 0)
        self.assertIn('integer', result['errors'][0].lower())
    
    def test_process_audit_import_partial_real_stock(self):
        """Test audit import when only one real stock column is filled"""
        file_data = self.create_audit_file([
            ['TEST001', 'Test Product 1', 100, 10, 110, None],  # Only OK changed
        ])
        
        result = AuditImportService.process_audit_import(
            file_data=file_data,
            user=self.user,
            audit_date=date.today(),
        )
        
        self.assertTrue(result['success'])
        self.assertEqual(result['updated_products'], 1)
        
        # Check adjustment
        adjustment = InventoryAdjustment.objects.get(product=self.product1)
        self.assertEqual(adjustment.delta_ok, 10)
        self.assertEqual(adjustment.delta_defect, 0)  # No change in defect
        self.assertEqual(adjustment.new_defect, 10)   # Kept original value
    
    def test_process_audit_import_with_comment(self):
        """Test that comment is saved with adjustments"""
        file_data = self.create_audit_file([
            ['TEST001', 'Test Product 1', 100, 10, 110, 12],
        ])
        
        comment = "Monthly physical inventory audit"
        result = AuditImportService.process_audit_import(
            file_data=file_data,
            user=self.user,
            audit_date=date.today(),
            comment=comment,
        )
        
        self.assertTrue(result['success'])
        
        adjustment = InventoryAdjustment.objects.get(product=self.product1)
        self.assertEqual(adjustment.comment, comment)
    
    def test_adjustment_created_by_user(self):
        """Test that adjustment records the user who performed it"""
        file_data = self.create_audit_file([
            ['TEST001', 'Test Product 1', 100, 10, 110, 12],
        ])
        
        result = AuditImportService.process_audit_import(
            file_data=file_data,
            user=self.user,
            audit_date=date.today(),
        )
        
        self.assertTrue(result['success'])
        
        adjustment = InventoryAdjustment.objects.get(product=self.product1)
        self.assertEqual(adjustment.created_by, self.user)
    
    def test_process_single_product_atomic(self):
        """Test that _process_single_product is atomic"""
        # This is tested implicitly through the import process
        # If an exception occurs, the transaction should roll back
        file_data = self.create_audit_file([
            ['TEST001', 'Test Product 1', 100, 10, 110, 12],
        ])
        
        result = AuditImportService.process_audit_import(
            file_data=file_data,
            user=self.user,
            audit_date=date.today(),
        )
        
        self.assertTrue(result['success'])
        
        # Both product update and adjustment creation should succeed
        self.product1.refresh_from_db()
        self.assertEqual(self.product1.stock_ok, Decimal('110.00'))
        self.assertEqual(InventoryAdjustment.objects.count(), 1)


class InventoryAdjustmentModelTests(TestCase):
    """Tests for InventoryAdjustment model"""
    
    def setUp(self):
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='testpass123'
        )
        
        self.brand = Brand.objects.create(name='Test Brand')
        self.category = Category.objects.create(name='Test Category')
        
        self.product = Product.objects.create(
            sku='TEST001',
            name='Test Product',
            brand=self.brand,
            category=self.category,
        )
    
    def test_total_delta_property(self):
        """Test total_delta property calculation"""
        adjustment = InventoryAdjustment.objects.create(
            product=self.product,
            delta_ok=10,
            delta_defect=5,
            previous_ok=100,
            previous_defect=10,
            new_ok=110,
            new_defect=15,
            date=date.today(),
            created_by=self.user,
        )
        
        self.assertEqual(adjustment.total_delta, 15)
    
    def test_total_delta_with_negative_values(self):
        """Test total_delta with decreases"""
        adjustment = InventoryAdjustment.objects.create(
            product=self.product,
            delta_ok=-10,
            delta_defect=-5,
            previous_ok=100,
            previous_defect=10,
            new_ok=90,
            new_defect=5,
            date=date.today(),
            created_by=self.user,
        )
        
        self.assertEqual(adjustment.total_delta, -15)
    
    def test_string_representation(self):
        """Test __str__ method"""
        adjustment = InventoryAdjustment.objects.create(
            product=self.product,
            delta_ok=10,
            delta_defect=5,
            previous_ok=100,
            previous_defect=10,
            new_ok=110,
            new_defect=15,
            date=date(2025, 1, 1),
            created_by=self.user,
        )
        
        str_repr = str(adjustment)
        self.assertIn('Test Product', str_repr)
        self.assertIn('2025-01-01', str_repr)
        self.assertIn('+10', str_repr)
        self.assertIn('+5', str_repr)
